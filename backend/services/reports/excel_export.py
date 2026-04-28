"""
Excel Report Generator — Multi-sheet financial & CRM reports.
Uses openpyxl to create professional Excel workbooks.
"""
import io
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from sqlalchemy.orm import Session
from database.models import (
    Client, Staff, VisitHistory, Appointment, Expense, Invoice, InvoiceItem,
    StaffPayment, StaffCommission, Tenant, Service, StaffServiceCommission,
    Product, InventoryMovement,
)

# Styles
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
MONEY_FORMAT = '#,##0'
DATE_FORMAT = 'DD/MM/YYYY'
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)


def _style_header(ws, row=1, cols=1):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_finance_report(db: Session, tenant_id: int, date_from: date, date_to: date) -> io.BytesIO:
    """Generate multi-sheet finance Excel report."""
    wb = Workbook()

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    biz_name = tenant.name if tenant else "Negocio"

    # ===== SHEET 1: RESUMEN =====
    ws = wb.active
    ws.title = "Resumen"
    ws.append([f"Reporte Financiero — {biz_name}"])
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='1E40AF')
    ws.append([f"Periodo: {date_from.strftime('%d/%m/%Y')} — {date_to.strftime('%d/%m/%Y')}"])
    ws['A2'].font = Font(name='Calibri', size=11, color='666666')
    ws.append([])

    visits = db.query(VisitHistory).filter(
        VisitHistory.status == "completed",
        VisitHistory.visit_date >= date_from,
        VisitHistory.visit_date <= date_to,
    )
    if tenant_id:
        visits = visits.filter(VisitHistory.tenant_id == tenant_id)
    visits = visits.all()

    total_revenue = sum(v.amount or 0 for v in visits)
    total_visits = len(visits)

    expenses = db.query(Expense).filter(
        Expense.date >= date_from, Expense.date <= date_to,
    )
    if tenant_id:
        expenses = expenses.filter(Expense.tenant_id == tenant_id)
    expenses = expenses.all()
    total_expenses = sum(e.amount or 0 for e in expenses)

    ws.append(["Metrica", "Valor"])
    _style_header(ws, row=4, cols=2)
    ws.append(["Ingresos totales", total_revenue])
    ws.append(["Gastos totales", total_expenses])
    ws.append(["Ganancia bruta", total_revenue - total_expenses])
    ws.append(["Total servicios realizados", total_visits])
    ws.append(["Ticket promedio", round(total_revenue / total_visits) if total_visits else 0])

    for row in range(5, 10):
        ws.cell(row=row, column=2).number_format = MONEY_FORMAT

    # ===== SHEET 2: INGRESOS (detalle) =====
    ws2 = wb.create_sheet("Ingresos")
    headers = ["Fecha", "Cliente", "Servicio", "Profesional", "Monto", "Metodo de Pago"]
    ws2.append(headers)
    _style_header(ws2, row=1, cols=len(headers))

    for v in sorted(visits, key=lambda x: x.visit_date, reverse=True):
        client = db.query(Client).filter(Client.id == v.client_id).first()
        staff = db.query(Staff).filter(Staff.id == v.staff_id).first()
        ws2.append([
            v.visit_date.strftime('%d/%m/%Y') if v.visit_date else '',
            client.name if client else 'Desconocido',
            v.service_name or '',
            staff.name if staff else 'Desconocido',
            v.amount or 0,
            v.payment_method or 'Sin registrar',
        ])
    for row in ws2.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _auto_width(ws2)

    # ===== SHEET 3: GASTOS =====
    ws3 = wb.create_sheet("Gastos")
    headers = ["Fecha", "Categoria", "Subcategoria", "Proveedor", "Descripcion", "Monto", "Metodo"]
    ws3.append(headers)
    _style_header(ws3, row=1, cols=len(headers))

    for e in sorted(expenses, key=lambda x: x.date or date.min, reverse=True):
        ws3.append([
            e.date.strftime('%d/%m/%Y') if e.date else '',
            e.category or '',
            e.subcategory or '',
            e.vendor or '',
            e.description or '',
            e.amount or 0,
            e.payment_method or '',
        ])
    for row in ws3.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _auto_width(ws3)

    # ===== SHEET 4: COMISIONES POR STAFF =====
    ws4 = wb.create_sheet("Comisiones")
    headers = ["Profesional", "Servicios", "Ingresos", "Tasa Comision", "Comision", "Propinas"]
    ws4.append(headers)
    _style_header(ws4, row=1, cols=len(headers))

    staff_data = {}
    for v in visits:
        sid = v.staff_id
        if sid not in staff_data:
            staff = db.query(Staff).filter(Staff.id == sid).first()
            comm = db.query(StaffCommission).filter(StaffCommission.staff_id == sid).first()
            rate = comm.default_rate if comm else 0.4
            staff_data[sid] = {
                'name': staff.name if staff else 'Desconocido',
                'count': 0, 'revenue': 0, 'rate': rate, 'tips': 0,
            }
        staff_data[sid]['count'] += 1
        staff_data[sid]['revenue'] += (v.amount or 0)
        staff_data[sid]['tips'] += (v.tip_amount or 0) if hasattr(v, 'tip_amount') else 0

    for sid, d in sorted(staff_data.items(), key=lambda x: -x[1]['revenue']):
        ws4.append([
            d['name'], d['count'], d['revenue'],
            f"{d['rate']*100:.0f}%",
            round(d['revenue'] * d['rate']),
            d['tips'],
        ])
    for row in ws4.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    for row in ws4.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _auto_width(ws4)

    # ===== SHEET 5: NOMINA =====
    ws5 = wb.create_sheet("Nomina")
    headers = ["Fecha", "Profesional", "Concepto", "Monto", "Metodo", "Referencia"]
    ws5.append(headers)
    _style_header(ws5, row=1, cols=len(headers))

    payments = db.query(StaffPayment).filter(
        StaffPayment.paid_at >= datetime.combine(date_from, datetime.min.time()),
        StaffPayment.paid_at <= datetime.combine(date_to, datetime.max.time()),
    )
    if tenant_id:
        payments = payments.filter(StaffPayment.tenant_id == tenant_id)

    for p in payments.order_by(StaffPayment.paid_at.desc()).all():
        staff = db.query(Staff).filter(Staff.id == p.staff_id).first()
        ws5.append([
            p.paid_at.strftime('%d/%m/%Y') if p.paid_at else '',
            staff.name if staff else 'Desconocido',
            p.concept or '',
            p.amount or 0,
            p.payment_method or '',
            p.reference or '',
        ])
    for row in ws5.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _auto_width(ws5)

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_clients_report(db: Session, tenant_id: int) -> io.BytesIO:
    """Generate Excel report of all clients with full data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = [
        "Codigo", "Nombre", "Telefono", "Email", "Cumpleanos",
        "Estado", "Visitas", "Total Gastado", "Ticket Promedio",
        "Ultima Visita", "Staff Preferido", "Tags",
    ]
    ws.append(headers)
    _style_header(ws, row=1, cols=len(headers))

    clients = db.query(Client).filter(Client.is_active == True)
    if tenant_id:
        clients = clients.filter(Client.tenant_id == tenant_id)

    from routes._helpers import compute_client_fields
    for c in clients.order_by(Client.name).all():
        try:
            fields = compute_client_fields(c, db)
        except Exception:
            fields = type('obj', (object,), {
                'total_visits': 0, 'total_spent': 0, 'avg_ticket': 0,
                'last_visit_date': None, 'status': 'nuevo',
            })()

        pref_staff = db.query(Staff).filter(Staff.id == c.preferred_barber_id).first() if c.preferred_barber_id else None

        ws.append([
            c.client_id or '',
            c.name or '',
            c.phone or '',
            c.email or '',
            c.birthday.strftime('%d/%m/%Y') if c.birthday else '',
            getattr(fields, 'status', c.status_override or 'nuevo'),
            getattr(fields, 'total_visits', 0),
            getattr(fields, 'total_spent', 0),
            getattr(fields, 'avg_ticket', 0),
            getattr(fields, 'last_visit_date', ''),
            pref_staff.name if pref_staff else '',
            ', '.join(c.tags) if c.tags else '',
        ])

    for row in ws.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_services_report(db: Session, tenant_id: int) -> io.BytesIO:
    """Premium services export — catalogo + comisiones por staff en hojas separadas."""
    wb = Workbook()

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    biz_name = tenant.name if tenant else "Negocio"

    # ===== SHEET 1: CATALOGO =====
    ws = wb.active
    ws.title = "Catalogo"
    ws.append([f"Catalogo de Servicios — {biz_name}"])
    ws.merge_cells('A1:K1')
    ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='1E40AF')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws.merge_cells('A2:K2')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='64748B')
    ws.append([])

    headers = [
        "ID", "Nombre", "Categoria", "Tipo", "Precio", "Duracion (min)",
        "Estado", "Modo Lina", "# Personal", "Personal asignado", "Descripcion",
    ]
    ws.append(headers)
    _style_header(ws, row=4, cols=len(headers))

    services = db.query(Service)
    if tenant_id:
        services = services.filter(Service.tenant_id == tenant_id)
    services = services.order_by(Service.category, Service.name).all()

    # Pre-fetch staff names
    all_staff_ids = set()
    for s in services:
        all_staff_ids.update(s.staff_ids or [])
    staff_map = {st.id: st for st in db.query(Staff).filter(Staff.id.in_(all_staff_ids)).all()} if all_staff_ids else {}

    for svc in services:
        ids = svc.staff_ids or []
        names = ", ".join([staff_map[i].name for i in ids if i in staff_map])
        ws.append([
            svc.id,
            svc.name or '',
            svc.category or '',
            (svc.service_type or 'cita').capitalize(),
            svc.price or 0,
            svc.duration_minutes or 0,
            'Activo' if svc.is_active else 'Inactivo',
            (svc.ai_mode or 'auto').capitalize(),
            len(ids),
            names,
            (svc.description or '')[:500],
        ])

    # Money column
    for row in ws.iter_rows(min_row=5, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _auto_width(ws)

    # ===== SHEET 2: COMISIONES POR STAFF =====
    ws2 = wb.create_sheet("Comisiones")
    ws2.append([f"Comisiones por servicio y profesional — {biz_name}"])
    ws2.merge_cells('A1:H1')
    ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='1E40AF')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.append([])

    headers2 = ["Servicio", "Categoria", "Precio", "Profesional", "Activo", "Tipo", "% / $", "Ganancia"]
    ws2.append(headers2)
    _style_header(ws2, row=3, cols=len(headers2))

    comm_q = db.query(StaffServiceCommission)
    if tenant_id:
        comm_q = comm_q.filter(StaffServiceCommission.tenant_id == tenant_id)
    all_comms = comm_q.all()
    comm_map = {(c.service_id, c.staff_id): c for c in all_comms}

    all_staff_q = db.query(Staff)
    if tenant_id:
        all_staff_q = all_staff_q.filter(Staff.tenant_id == tenant_id)
    all_staff = all_staff_q.order_by(Staff.name).all()

    for svc in services:
        for st in all_staff:
            c = comm_map.get((svc.id, st.id))
            if not c:
                continue
            ctype = c.commission_type or 'percentage'
            value = (
                f"${c.commission_amount or 0}" if ctype == 'fixed'
                else f"{int((c.commission_rate or 0) * 100)}%"
            )
            earnings = (
                int(c.commission_amount or 0) if ctype == 'fixed'
                else int((svc.price or 0) * (c.commission_rate or 0))
            )
            ws2.append([
                svc.name or '',
                svc.category or '',
                svc.price or 0,
                st.name or '',
                'Si' if c.is_enabled else 'No',
                'Fijo' if ctype == 'fixed' else 'Porcentaje',
                value,
                earnings,
            ])

    for row in ws2.iter_rows(min_row=4, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    for row in ws2.iter_rows(min_row=4, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _auto_width(ws2)

    # ===== SHEET 3: PLANTILLA IMPORTACION =====
    ws3 = wb.create_sheet("Plantilla Import")
    ws3.append([
        "Para crear servicios masivamente: completa esta hoja y subela en 'Importar'. "
        "Las columnas obligatorias son: Nombre, Categoria, Precio."
    ])
    ws3.merge_cells('A1:H1')
    ws3['A1'].font = Font(name='Calibri', italic=True, size=10, color='64748B')
    ws3['A1'].alignment = Alignment(horizontal='left', wrap_text=True)
    ws3.row_dimensions[1].height = 30
    ws3.append([])

    template_headers = [
        "Nombre*", "Categoria*", "Precio*", "Duracion (min)",
        "Tipo (cita|paquete|reserva)", "Estado (activo|inactivo)",
        "Modo Lina (auto|manual)", "Descripcion",
    ]
    ws3.append(template_headers)
    _style_header(ws3, row=3, cols=len(template_headers))
    ws3.append(["Corte cabello", "Barberia", 30000, 30, "cita", "activo", "auto", "Corte clasico de caballero"])
    ws3.append(["Manicure semipermanente", "Arte en Unas", 45000, 60, "cita", "activo", "auto", ""])
    _auto_width(ws3)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_inventory_report(db: Session, tenant_id: int) -> io.BytesIO:
    """Premium inventory export — productos + movimientos + plantilla."""
    wb = Workbook()

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    biz_name = tenant.name if tenant else 'Negocio'

    # ===== SHEET 1: PRODUCTOS =====
    ws = wb.active
    ws.title = 'Productos'
    ws.append([f'Inventario — {biz_name}'])
    ws.merge_cells('A1:K1')
    ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='1E40AF')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
    ws.merge_cells('A2:K2')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='64748B')
    ws.append([])

    headers = ['ID', 'Nombre', 'SKU / Codigo', 'Categoria', 'Precio venta', 'Costo', 'Margen %', 'Stock', 'Min stock', 'Estado', 'Proveedor']
    ws.append(headers)
    _style_header(ws, row=4, cols=len(headers))

    products = db.query(Product)
    if tenant_id:
        products = products.filter(Product.tenant_id == tenant_id)
    products = products.order_by(Product.category, Product.name).all()

    for p in products:
        margin = 0
        if (p.price or 0) > 0:
            margin = round((((p.price or 0) - (p.cost or 0)) / (p.price or 1)) * 100, 1)
        ws.append([
            p.id,
            p.name or '',
            p.sku or '',
            p.category or '',
            p.price or 0,
            p.cost or 0,
            margin,
            p.stock or 0,
            p.min_stock or 0,
            'Activo' if p.is_active else 'Inactivo',
            p.supplier or '',
        ])

    # Money columns
    for row in ws.iter_rows(min_row=5, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _auto_width(ws)

    # ===== SHEET 2: MOVIMIENTOS =====
    ws2 = wb.create_sheet('Movimientos')
    ws2.append([f'Movimientos de inventario — {biz_name}'])
    ws2.merge_cells('A1:G1')
    ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='1E40AF')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.append([])

    headers2 = ['Fecha', 'Producto', 'Tipo', 'Cantidad', 'Costo unidad', 'Usuario', 'Nota']
    ws2.append(headers2)
    _style_header(ws2, row=3, cols=len(headers2))

    mvs = db.query(InventoryMovement).filter(InventoryMovement.tenant_id == tenant_id).order_by(InventoryMovement.created_at.desc()).limit(2000).all() if tenant_id else []
    prod_map = {p.id: p for p in products}
    for m in mvs:
        prod = prod_map.get(m.product_id)
        ws2.append([
            m.created_at.strftime('%d/%m/%Y %H:%M') if m.created_at else '',
            prod.name if prod else f'#{m.product_id}',
            (m.movement_type or '').capitalize(),
            m.quantity or 0,
            m.unit_cost or 0,
            m.created_by or '',
            (m.note or '')[:300],
        ])

    for row in ws2.iter_rows(min_row=4, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _auto_width(ws2)

    # ===== SHEET 3: PLANTILLA IMPORT =====
    ws3 = wb.create_sheet('Plantilla Import')
    ws3.append(['Para crear productos masivamente: completa esta hoja y subela en Importar. Obligatorios: Nombre, Precio venta, Costo.'])
    ws3.merge_cells('A1:H1')
    ws3['A1'].font = Font(name='Calibri', italic=True, size=10, color='64748B')
    ws3['A1'].alignment = Alignment(horizontal='left', wrap_text=True)
    ws3.row_dimensions[1].height = 30
    ws3.append([])

    template_headers = ['Nombre*', 'Categoria', 'SKU / Codigo', 'Precio venta*', 'Costo*', 'Stock inicial', 'Min stock', 'Proveedor']
    ws3.append(template_headers)
    _style_header(ws3, row=3, cols=len(template_headers))
    ws3.append(['Shampoo profesional 250ml', 'Cuidado capilar', 'SHA-001', 35000, 18000, 10, 3, 'Distribuidora XYZ'])
    ws3.append(['Cera para cabello 100ml', 'Styling', 'WAX-002', 28000, 12000, 6, 2, ''])
    _auto_width(ws3)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

