"""
Plexify Studio — Inventory & Product Management Endpoints
CRUD products, stock movements, low-stock alerts, reports.
Multi-tenant isolated via safe_tid().
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
from typing import Optional

from database.connection import get_db
from database.models import Product, InventoryMovement, Admin, Checkout, CheckoutItem
from routes._helpers import safe_tid


def _get_checkout_info(db, checkout_id):
    """Get client/staff info from a checkout for movement history."""
    try:
        co = db.query(Checkout).filter(Checkout.id == checkout_id).first()
        if not co:
            return {}
        # Find which staff sold this product
        prod_items = db.query(CheckoutItem).filter(
            CheckoutItem.checkout_id == checkout_id,
            CheckoutItem.service_name.ilike('[Producto]%')
        ).all()
        staff_name = prod_items[0].staff_name if prod_items else co.staff_name
        return {
            "client_name": co.client_name,
            "client_phone": getattr(co, 'client_phone', None),
            "staff_name": staff_name,
            "payment_method": co.payment_method,
        }
    except Exception:
        return {}
from middleware.auth_middleware import get_current_user

router = APIRouter(prefix="/inventory", tags=["Inventory"])


# ============================================================================
# IMPORT / EXPORT — must be BEFORE /products/{id} routes
# ============================================================================

@router.get("/export")
def export_products_xlsx(db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    """Export full inventory + movements + plantilla as styled xlsx."""
    from fastapi.responses import StreamingResponse
    from services.reports.excel_export import generate_inventory_report
    tid = safe_tid(user, db)
    if not tid:
        raise HTTPException(400, "No tenant assigned")
    buf = generate_inventory_report(db, tid)
    filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import")
async def import_products_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    """Bulk-create products from xlsx/csv. Skips duplicates by name+sku."""
    import io as _io
    import csv as _csv
    tid = safe_tid(user, db)
    if not tid:
        raise HTTPException(400, "No tenant assigned")
    if not file.filename:
        raise HTTPException(400, "Archivo requerido")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    content = await file.read()

    rows = []
    if ext == "csv":
        text = content.decode("utf-8-sig")
        rows = list(_csv.DictReader(_io.StringIO(text)))
    elif ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
            ws = None
            for name in ("Plantilla Import", "Productos"):
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            if ws is None:
                ws = wb.active
            header_row_idx = 1
            headers = []
            for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                non_empty = [c for c in row if c is not None and str(c).strip()]
                if len(non_empty) >= 3 and any('nombre' in str(c).lower() for c in row if c):
                    header_row_idx = ridx
                    headers = [str(c or '').strip() for c in row]
                    break
            if not headers:
                raise HTTPException(400, "No se encontró fila de encabezados")
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if all(c is None or not str(c).strip() for c in row):
                    continue
                d = {}
                for i, c in enumerate(row):
                    if i < len(headers) and headers[i]:
                        d[headers[i]] = c
                rows.append(d)
        except ImportError:
            raise HTTPException(400, "Instala openpyxl para importar Excel")
    else:
        raise HTTPException(400, "Formato no soportado. Usa CSV o XLSX")

    def gv(row, *keys):
        for k in keys:
            for rk in row:
                if rk and str(rk).lower().strip().rstrip('*') == k.lower():
                    v = row[rk]
                    if v is None:
                        return ""
                    return str(v).strip()
        return ""

    existing = {(p.name.lower().strip(), (p.sku or '').lower().strip()): p.id
                for p in db.query(Product).filter(Product.tenant_id == tid).all()}

    imported, skipped, errors = 0, 0, []
    for i, row in enumerate(rows):
        line_no = i + 2
        try:
            name = gv(row, "nombre", "name")
            sku = gv(row, "sku / codigo", "sku", "codigo", "code")
            price_raw = gv(row, "precio venta", "precio", "price", "venta")
            cost_raw = gv(row, "costo", "cost")
            if not name:
                errors.append({"row": line_no, "reason": "Nombre requerido"})
                continue
            try:
                price = float(str(price_raw).replace('$', '').replace(',', '').strip() or 0)
            except Exception:
                errors.append({"row": line_no, "name": name, "reason": "Precio inválido"})
                continue
            try:
                cost = float(str(cost_raw).replace('$', '').replace(',', '').strip() or 0)
            except Exception:
                cost = 0
            key = (name.lower().strip(), sku.lower().strip())
            if key in existing:
                skipped += 1
                continue
            category = gv(row, "categoria", "categoría", "category") or None
            stock_raw = gv(row, "stock inicial", "stock")
            try:
                stock = int(float(stock_raw)) if stock_raw else 0
            except Exception:
                stock = 0
            min_stock_raw = gv(row, "min stock", "minimo", "mínimo")
            try:
                min_stock = int(float(min_stock_raw)) if min_stock_raw else 5
            except Exception:
                min_stock = 5
            supplier = gv(row, "proveedor", "supplier") or None

            db.add(Product(
                tenant_id=tid,
                name=name,
                sku=sku or None,
                category=category,
                price=price,
                cost=cost,
                stock=stock,
                min_stock=min_stock,
                supplier=supplier,
                is_active=True,
            ))
            imported += 1
            existing[key] = -1
        except Exception as e:
            errors.append({"row": line_no, "reason": str(e)[:200]})

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors, "total_rows": len(rows)}


# ============================================================================
# PRODUCTS — CRUD
# ============================================================================

@router.get("/products")
def list_products(
    category: Optional[str] = None,
    low_stock: bool = False,
    active_only: bool = True,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    user: Admin = Depends(get_current_user),
):
    tid = safe_tid(user, db)
    q = db.query(Product)
    if tid:
        q = q.filter(Product.tenant_id == tid)
    if active_only:
        q = q.filter(Product.is_active == True)
    if category:
        q = q.filter(Product.category == category)
    if search:
        q = q.filter(Product.name.ilike(f"%{search}%"))
    if low_stock:
        q = q.filter(Product.stock <= Product.min_stock)

    products = q.order_by(Product.category, Product.name).all()

    # Calculate totals
    total_value = sum(p.stock * p.cost for p in products)
    total_retail = sum(p.stock * p.price for p in products)
    low_stock_count = sum(1 for p in products if p.stock <= p.min_stock)

    return {
        "products": [
            {
                "id": p.id,
                "name": p.name,
                "sku": p.sku,
                "category": p.category,
                "description": p.description,
                "price": p.price,
                "cost": p.cost,
                "margin": round(p.price - p.cost, 0) if p.price and p.cost else 0,
                "margin_pct": round(((p.price - p.cost) / p.price * 100), 1) if p.price and p.price > 0 else 0,
                "stock": p.stock,
                "min_stock": p.min_stock,
                "is_low_stock": p.stock <= p.min_stock,
                "supplier": p.supplier,
                "image_url": p.image_url,
                "is_active": p.is_active,
                "created_at": p.created_at.isoformat() if p.created_at else None,
            }
            for p in products
        ],
        "summary": {
            "total_products": len(products),
            "total_stock_value": total_value,
            "total_retail_value": total_retail,
            "potential_profit": total_retail - total_value,
            "low_stock_count": low_stock_count,
        },
    }


@router.get("/products/{product_id}")
def get_product(product_id: int, db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    tid = safe_tid(user, db)
    q = db.query(Product).filter(Product.id == product_id)
    if tid:
        q = q.filter(Product.tenant_id == tid)
    product = q.first()
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    # Get recent movements from inventory
    movements = (
        db.query(InventoryMovement)
        .filter(InventoryMovement.product_id == product_id)
        .order_by(InventoryMovement.created_at.desc())
        .limit(50)
        .all()
    )

    # Also get sales from CheckoutItem (for older sales without InventoryMovement)
    product_tag = f"[Producto] {product.name}"
    checkout_sales = (
        db.query(CheckoutItem, Checkout)
        .join(Checkout, Checkout.id == CheckoutItem.checkout_id)
        .filter(CheckoutItem.service_name == product_tag)
        .order_by(Checkout.created_at.desc())
        .limit(50)
        .all()
    )

    # Merge: existing movement IDs from checkout references
    existing_refs = {m.reference_id for m in movements if m.reference_id and m.movement_type == 'sale'}

    # Build combined list
    all_movements = []
    for m in movements:
        entry = {
            "id": m.id,
            "type": m.movement_type,
            "quantity": m.quantity,
            "unit_cost": m.unit_cost,
            "note": m.note,
            "reference_id": m.reference_id,
            "created_by": m.created_by,
            "created_at": m.created_at.isoformat() if m.created_at else None,
        }
        if m.movement_type == "sale" and m.reference_id:
            entry.update(_get_checkout_info(db, m.reference_id))
        all_movements.append(entry)

    # Add checkout sales that don't have InventoryMovement yet
    for ci, co in checkout_sales:
        if co.id in existing_refs:
            continue
        all_movements.append({
            "id": f"co-{co.id}",
            "type": "sale",
            "quantity": -(ci.quantity or 1),
            "unit_cost": ci.unit_price,
            "note": f"Venta checkout #{co.id}",
            "reference_id": co.id,
            "created_by": co.created_by,
            "created_at": co.created_at.isoformat() if co.created_at else None,
            "client_name": co.client_name,
            "client_phone": getattr(co, 'client_phone', None),
            "staff_name": ci.staff_name or co.staff_name,
            "payment_method": co.payment_method,
        })

    # Sort by date desc
    all_movements.sort(key=lambda x: x.get('created_at') or '', reverse=True)

    return {
        "id": product.id,
        "name": product.name,
        "sku": product.sku,
        "category": product.category,
        "description": product.description,
        "price": product.price,
        "cost": product.cost,
        "stock": product.stock,
        "min_stock": product.min_stock,
        "supplier": product.supplier,
        "image_url": product.image_url,
        "is_active": product.is_active,
        "movements": all_movements[:50],
    }


@router.post("/products")
def create_product(data: dict, db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    tid = safe_tid(user, db)
    if not tid:
        raise HTTPException(status_code=400, detail="No tenant asociado")

    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")

    product = Product(
        tenant_id=tid,
        name=name,
        sku=(data.get("sku") or "").strip() or None,
        category=(data.get("category") or "").strip() or None,
        description=(data.get("description") or "").strip() or None,
        price=float(data.get("price", 0)),
        cost=float(data.get("cost", 0)),
        stock=int(data.get("stock", 0)),
        min_stock=int(data.get("min_stock", 5)),
        supplier=(data.get("supplier") or "").strip() or None,
        image_url=(data.get("image_url") or "").strip() or None,
    )
    db.add(product)
    db.commit()
    db.refresh(product)

    # If initial stock > 0, create a purchase movement
    if product.stock > 0:
        movement = InventoryMovement(
            tenant_id=tid,
            product_id=product.id,
            movement_type="purchase",
            quantity=product.stock,
            unit_cost=product.cost,
            note="Stock inicial",
            created_by=getattr(user, 'username', 'admin'),
        )
        db.add(movement)
        db.commit()

    return {"id": product.id, "name": product.name}


@router.put("/products/{product_id}")
def update_product(product_id: int, data: dict, db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    tid = safe_tid(user, db)
    q = db.query(Product).filter(Product.id == product_id)
    if tid:
        q = q.filter(Product.tenant_id == tid)
    product = q.first()
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    for field in ["name", "sku", "category", "description", "supplier", "image_url"]:
        if field in data:
            setattr(product, field, (data[field] or "").strip() or None if isinstance(data[field], str) else data[field])

    for field in ["price", "cost"]:
        if field in data:
            setattr(product, field, float(data[field]))

    for field in ["min_stock"]:
        if field in data:
            setattr(product, field, int(data[field]))

    if "is_active" in data:
        product.is_active = bool(data["is_active"])

    db.commit()
    return {"id": product.id, "name": product.name}


@router.delete("/products/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    tid = safe_tid(user, db)
    q = db.query(Product).filter(Product.id == product_id)
    if tid:
        q = q.filter(Product.tenant_id == tid)
    product = q.first()
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    # Soft delete — keep history
    product.is_active = False
    db.commit()
    return {"success": True}


# ============================================================================
# STOCK MOVEMENTS
# ============================================================================

@router.post("/products/{product_id}/stock")
def adjust_stock(product_id: int, data: dict, db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    """Add stock (purchase), remove stock (adjustment/loss), or correct stock."""
    tid = safe_tid(user, db)
    q = db.query(Product).filter(Product.id == product_id)
    if tid:
        q = q.filter(Product.tenant_id == tid)
    product = q.first()
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    movement_type = data.get("type", "adjustment")  # purchase, adjustment, return, loss
    quantity = int(data.get("quantity", 0))
    if quantity == 0:
        raise HTTPException(status_code=400, detail="La cantidad no puede ser 0")

    unit_cost = data.get("unit_cost", product.cost)
    note = (data.get("note") or "").strip()

    # For losses/sales, quantity should be negative
    if movement_type in ("loss", "sale") and quantity > 0:
        quantity = -quantity

    # Validate stock won't go negative
    new_stock = product.stock + quantity
    if new_stock < 0:
        raise HTTPException(status_code=400, detail=f"Stock insuficiente. Disponible: {product.stock}")

    movement = InventoryMovement(
        tenant_id=tid or product.tenant_id,
        product_id=product.id,
        movement_type=movement_type,
        quantity=quantity,
        unit_cost=float(unit_cost) if unit_cost else None,
        note=note,
        created_by=getattr(user, 'username', 'admin'),
    )
    db.add(movement)

    # Update cost on purchase (weighted average)
    if movement_type == "purchase" and quantity > 0 and unit_cost:
        old_value = product.stock * product.cost
        new_value = quantity * float(unit_cost)
        total_units = product.stock + quantity
        if total_units > 0:
            product.cost = round((old_value + new_value) / total_units, 0)

    product.stock = new_stock
    db.commit()

    return {
        "product_id": product.id,
        "new_stock": product.stock,
        "movement_type": movement_type,
        "quantity": quantity,
        "is_low_stock": product.stock <= product.min_stock,
    }


@router.get("/movements")
def list_movements(
    product_id: Optional[int] = None,
    movement_type: Optional[str] = None,
    limit: int = 50,
    db: Session = Depends(get_db),
    user: Admin = Depends(get_current_user),
):
    tid = safe_tid(user, db)
    q = db.query(InventoryMovement)
    if tid:
        q = q.filter(InventoryMovement.tenant_id == tid)
    if product_id:
        q = q.filter(InventoryMovement.product_id == product_id)
    if movement_type:
        q = q.filter(InventoryMovement.movement_type == movement_type)

    movements = q.order_by(InventoryMovement.created_at.desc()).limit(limit).all()

    return [
        {
            "id": m.id,
            "product_id": m.product_id,
            "product_name": m.product.name if m.product else "?",
            "type": m.movement_type,
            "quantity": m.quantity,
            "unit_cost": m.unit_cost,
            "note": m.note,
            "created_by": m.created_by,
            "created_at": m.created_at.isoformat() if m.created_at else None,
        }
        for m in movements
    ]


# ============================================================================
# ALERTS & REPORTS
# ============================================================================

@router.get("/alerts")
def low_stock_alerts(db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    tid = safe_tid(user, db)
    q = db.query(Product).filter(
        Product.is_active == True,
        Product.stock <= Product.min_stock,
    )
    if tid:
        q = q.filter(Product.tenant_id == tid)

    products = q.order_by(Product.stock.asc()).all()
    return [
        {
            "id": p.id,
            "name": p.name,
            "stock": p.stock,
            "min_stock": p.min_stock,
            "category": p.category,
            "supplier": p.supplier,
        }
        for p in products
    ]


@router.get("/categories")
def list_categories(db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    tid = safe_tid(user, db)
    q = db.query(Product.category, func.count(Product.id).label("count")).filter(
        Product.is_active == True,
        Product.category.isnot(None),
    )
    if tid:
        q = q.filter(Product.tenant_id == tid)
    results = q.group_by(Product.category).order_by(Product.category).all()
    return [{"category": r[0], "count": r[1]} for r in results]


@router.get("/report")
def inventory_report(db: Session = Depends(get_db), user: Admin = Depends(get_current_user)):
    """Summary report: total value, margins, top sellers, slow movers."""
    tid = safe_tid(user, db)
    q = db.query(Product).filter(Product.is_active == True)
    if tid:
        q = q.filter(Product.tenant_id == tid)
    products = q.all()

    # Sales movements (last 30 days)
    from datetime import timedelta
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    sales_q = db.query(
        InventoryMovement.product_id,
        func.sum(func.abs(InventoryMovement.quantity)).label("sold"),
    ).filter(
        InventoryMovement.movement_type == "sale",
        InventoryMovement.created_at >= thirty_days_ago,
    )
    if tid:
        sales_q = sales_q.filter(InventoryMovement.tenant_id == tid)
    sales = {r[0]: r[1] for r in sales_q.group_by(InventoryMovement.product_id).all()}

    product_data = []
    for p in products:
        sold_30d = sales.get(p.id, 0)
        product_data.append({
            "id": p.id,
            "name": p.name,
            "category": p.category,
            "stock": p.stock,
            "price": p.price,
            "cost": p.cost,
            "stock_value": p.stock * p.cost,
            "retail_value": p.stock * p.price,
            "sold_30d": sold_30d,
            "revenue_30d": sold_30d * p.price,
            "margin_pct": round(((p.price - p.cost) / p.price * 100), 1) if p.price > 0 else 0,
        })

    # Sort by sales for top sellers
    top_sellers = sorted(product_data, key=lambda x: x["sold_30d"], reverse=True)[:10]
    slow_movers = sorted([p for p in product_data if p["sold_30d"] == 0], key=lambda x: x["stock"], reverse=True)[:10]

    total_stock_value = sum(p["stock_value"] for p in product_data)
    total_retail_value = sum(p["retail_value"] for p in product_data)
    total_revenue_30d = sum(p["revenue_30d"] for p in product_data)

    return {
        "total_products": len(products),
        "total_stock_value": total_stock_value,
        "total_retail_value": total_retail_value,
        "potential_profit": total_retail_value - total_stock_value,
        "revenue_30d": total_revenue_30d,
        "top_sellers": top_sellers,
        "slow_movers": slow_movers,
    }
