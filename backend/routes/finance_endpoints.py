from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date, timedelta
from typing import Optional
import csv
import io

from database.connection import get_db
from database.models import (
    Admin, Staff, Client, VisitHistory, Expense, Invoice, InvoiceItem,
    StaffCommission, StaffServiceCommission, Service, StaffFine,
    Checkout, Appointment, Product, InventoryMovement, CashMovement,
    Tenant,
)
from middleware.auth_middleware import get_current_user
from routes._helpers import safe_tid
from schemas import (
    ExpenseCreate, ExpenseUpdate, ExpenseResponse, ExpenseSummaryItem,
    CommissionConfigResponse, CommissionConfigUpdate, CommissionPayoutItem,
    InvoiceCreate, InvoiceUpdate, InvoiceResponse, InvoiceItemResponse,
    PnLResponse, PaymentMethodItem, ImportResult, ImportRowResult, UninvoicedVisitResponse,
)

router = APIRouter()


# ============================================================================
# HELPERS
# ============================================================================

def _parse_period(period: str, date_from: Optional[str], date_to: Optional[str]):
    today = datetime.utcnow().date()
    if date_from and date_to:
        return date.fromisoformat(date_from), date.fromisoformat(date_to)
    if period == "today":
        return today, today
    if period == "week":
        return today - timedelta(days=today.weekday()), today
    if period == "year":
        return date(today.year, 1, 1), today
    # month default
    return date(today.year, today.month, 1), today


def _tenant_filter(query, model, tid):
    """Apply tenant_id filter if the user belongs to a tenant."""
    if tid:
        return query.filter(model.tenant_id == tid)
    return query


# ============================================================================
# EXPENSES
# ============================================================================

@router.get("/expenses/")
def list_expenses(
    period: str = Query("month"),
    category: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    limit: Optional[int] = Query(None, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    start, end = _parse_period(period, date_from, date_to)
    q = db.query(Expense).filter(Expense.date >= start, Expense.date <= end, Expense.deleted_at.is_(None))
    q = _tenant_filter(q, Expense, tid)
    if category:
        q = q.filter(Expense.category == category)
    q = q.order_by(Expense.date.desc())
    if limit is not None:
        total = q.count()
        items = q.offset(offset).limit(limit).all()
        return {"items": items, "total": total, "limit": limit, "offset": offset}
    return q.all()


@router.post("/expenses/", response_model=ExpenseResponse)
def create_expense(data: ExpenseCreate, db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    tid = safe_tid(current_user, db)
    expense = Expense(tenant_id=tid, **data.model_dump())
    db.add(expense)
    db.commit()
    db.refresh(expense)
    return expense


@router.put("/expenses/{expense_id}", response_model=ExpenseResponse)
def update_expense(expense_id: int, data: ExpenseUpdate, db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    tid = safe_tid(current_user, db)
    q = db.query(Expense).filter(Expense.id == expense_id)
    q = _tenant_filter(q, Expense, tid)
    expense = q.first()
    if not expense:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(expense, k, v)
    db.commit()
    db.refresh(expense)
    return expense


@router.delete("/expenses/{expense_id}")
def delete_expense(expense_id: int, db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    tid = safe_tid(current_user, db)
    q = db.query(Expense).filter(Expense.id == expense_id)
    q = _tenant_filter(q, Expense, tid)
    expense = q.first()
    if not expense:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    expense.deleted_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


@router.get("/expenses/summary")
def expenses_summary(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    start, end = _parse_period(period, date_from, date_to)
    q = db.query(Expense).filter(Expense.date >= start, Expense.date <= end)
    q = _tenant_filter(q, Expense, tid)
    expenses = q.all()
    total = sum(e.amount for e in expenses)
    cat_map = {}
    for e in expenses:
        if e.category not in cat_map:
            cat_map[e.category] = {"total": 0, "count": 0}
        cat_map[e.category]["total"] += e.amount
        cat_map[e.category]["count"] += 1

    items = sorted(
        [
            ExpenseSummaryItem(
                category=c,
                total=info["total"],
                count=info["count"],
                pct_of_total=round((info["total"] / total * 100), 1) if total > 0 else 0,
            )
            for c, info in cat_map.items()
        ],
        key=lambda x: x.total,
        reverse=True,
    )
    return {"total": total, "count": len(expenses), "by_category": items}


# ============================================================================
# COMMISSIONS
# ============================================================================

@router.get("/finances/commissions/config", response_model=list[CommissionConfigResponse])
def list_commission_configs(db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    tid = safe_tid(current_user, db)
    q = db.query(Staff).filter(Staff.is_active == True)
    q = _tenant_filter(q, Staff, tid)
    staff_list = q.all()
    result = []
    for s in staff_list:
        comm = db.query(StaffCommission).filter(StaffCommission.staff_id == s.id).first()
        result.append(CommissionConfigResponse(
            staff_id=s.id,
            staff_name=s.name,
            default_rate=comm.default_rate if comm else 0.40,
            service_overrides=comm.service_overrides if comm else {},
        ))
    return result


@router.put("/finances/commissions/config/{staff_id}", response_model=CommissionConfigResponse)
def update_commission_config(staff_id: int, data: CommissionConfigUpdate, db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    tid = safe_tid(current_user, db)
    q = db.query(Staff).filter(Staff.id == staff_id)
    q = _tenant_filter(q, Staff, tid)
    staff = q.first()
    if not staff:
        raise HTTPException(status_code=404, detail="Profesional no encontrado")

    comm = db.query(StaffCommission).filter(StaffCommission.staff_id == staff_id).first()
    if not comm:
        comm = StaffCommission(tenant_id=tid, staff_id=staff_id)
        db.add(comm)

    comm.default_rate = data.default_rate
    comm.service_overrides = data.service_overrides
    db.commit()
    db.refresh(comm)

    return CommissionConfigResponse(
        staff_id=staff.id,
        staff_name=staff.name,
        default_rate=comm.default_rate,
        service_overrides=comm.service_overrides,
    )


@router.get("/finances/commissions/payouts", response_model=list[CommissionPayoutItem])
def get_commission_payouts(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    start, end = _parse_period(period, date_from, date_to)
    q = (
        db.query(VisitHistory)
        .filter(VisitHistory.status == "completed")
        .filter(VisitHistory.visit_date >= start, VisitHistory.visit_date <= end)
    )
    q = _tenant_filter(q, VisitHistory, tid)
    visits = q.all()

    # Also get appointments with locked commissions in period
    from database.models import StaffServiceCommission
    apt_q = (
        db.query(Appointment)
        .filter(Appointment.status.in_(["completed", "paid"]))
        .filter(Appointment.date >= start, Appointment.date <= end)
    )
    apt_q = _tenant_filter(apt_q, Appointment, tid)
    paid_apts = apt_q.all()

    # Group by staff — use locked commission from appointments, fallback to per-service config
    staff_rev = {}
    for apt in paid_apts:
        sid = apt.staff_id
        if sid not in staff_rev:
            staff_rev[sid] = {"revenue": 0, "count": 0, "commission": 0}
        staff_rev[sid]["revenue"] += apt.price or 0
        staff_rev[sid]["count"] += 1

        if getattr(apt, 'commission_amount', None) is not None:
            # Use locked commission from payment time
            staff_rev[sid]["commission"] += apt.commission_amount
        elif getattr(apt, 'commission_rate', None) is not None:
            staff_rev[sid]["commission"] += int((apt.price or 0) * apt.commission_rate)
        else:
            # Lookup current per-service rate
            svc_comm = db.query(StaffServiceCommission).filter(
                StaffServiceCommission.staff_id == sid,
                StaffServiceCommission.service_id == apt.service_id,
            ).first()
            if svc_comm:
                staff_rev[sid]["commission"] += int((apt.price or 0) * svc_comm.commission_rate)
            else:
                # Fallback to old default_rate
                comm = db.query(StaffCommission).filter(StaffCommission.staff_id == sid).first()
                rate = comm.default_rate if comm else 0.0
                staff_rev[sid]["commission"] += int((apt.price or 0) * rate)

    # Also include visit history entries not covered by appointments
    for v in visits:
        if v.staff_id not in staff_rev:
            staff_rev[v.staff_id] = {"revenue": 0, "count": 0, "commission": 0}
            staff_rev[v.staff_id]["revenue"] += v.amount
            staff_rev[v.staff_id]["count"] += 1
            comm = db.query(StaffCommission).filter(StaffCommission.staff_id == v.staff_id).first()
            rate = comm.default_rate if comm else 0.0
            staff_rev[v.staff_id]["commission"] += int(v.amount * rate)

    result = []
    for sid, info in staff_rev.items():
        staff = db.query(Staff).filter(Staff.id == sid).first()
        avg_rate = info["commission"] / info["revenue"] if info["revenue"] > 0 else 0.0
        result.append(CommissionPayoutItem(
            staff_id=sid,
            staff_name=staff.name if staff else "Desconocido",
            rate=round(avg_rate, 4),
            total_revenue=info["revenue"],
            commission_amount=info["commission"],
            services_count=info["count"],
        ))

    return sorted(result, key=lambda x: x.commission_amount, reverse=True)


# ============================================================================
# UNINVOICED VISITS
# ============================================================================

@router.get("/finances/uninvoiced-visits", response_model=list[UninvoicedVisitResponse])
def get_uninvoiced_visits(
    client_id: Optional[int] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    """List completed visits that haven't been invoiced yet."""
    tid = safe_tid(current_user, db)
    q = (
        db.query(VisitHistory)
        .filter(VisitHistory.status == "completed")
        .filter(
            (VisitHistory.is_invoiced == False) | (VisitHistory.is_invoiced == None)
        )
    )
    q = _tenant_filter(q, VisitHistory, tid)
    if client_id:
        q = q.filter(VisitHistory.client_id == client_id)
    if date_from:
        q = q.filter(VisitHistory.visit_date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(VisitHistory.visit_date <= date.fromisoformat(date_to))

    visits = q.order_by(VisitHistory.visit_date.desc()).limit(50).all()

    result = []
    for v in visits:
        client = db.query(Client).filter(Client.id == v.client_id).first()
        staff = db.query(Staff).filter(Staff.id == v.staff_id).first()
        result.append(UninvoicedVisitResponse(
            id=v.id,
            client_id=v.client_id,
            client_name=client.name if client else "?",
            staff_name=staff.name if staff else "?",
            service_name=v.service_name,
            amount=v.amount,
            visit_date=v.visit_date,
        ))
    return result


# ============================================================================
# INVOICES
# ============================================================================

@router.get("/invoices/")
def list_invoices(
    status: Optional[str] = Query(None),
    limit: Optional[int] = Query(None, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    q = db.query(Invoice).filter(Invoice.deleted_at.is_(None))
    q = _tenant_filter(q, Invoice, tid)
    if status:
        q = q.filter(Invoice.status == status)
    q = q.order_by(Invoice.issued_date.desc())
    total_count = q.count() if limit is not None else None
    if limit is not None:
        invoices = q.offset(offset).limit(limit).all()
    else:
        invoices = q.all()

    # Enrich with tip + real commission from linked checkout
    from database.models import Checkout, StaffCommission
    result = []
    for inv in invoices:
        data = InvoiceResponse.model_validate(inv).model_dump()

        # Get tip from linked checkout
        checkout = db.query(Checkout).filter(Checkout.invoice_id == inv.id).first()
        data["tip"] = checkout.tip if checkout else 0
        if not data.get("discount_amount"):
            data["discount_amount"] = checkout.discount_amount if checkout else 0
        if not data.get("discount_type"):
            data["discount_type"] = checkout.discount_type if checkout else None
        data["receipt_url"] = checkout.receipt_url if checkout and getattr(checkout, 'receipt_url', None) else None
        data["payment_details"] = checkout.payment_details if checkout and getattr(checkout, 'payment_details', None) else None

        # Get commission — prefer FROZEN rates from CheckoutItem, fallback to current rates
        from database.models import CheckoutItem, StaffServiceCommission
        staff_name = None
        commission_rate = 0.5  # ultimate fallback
        frozen_items = []
        if checkout:
            ci_list = db.query(CheckoutItem).filter(CheckoutItem.checkout_id == checkout.id).all()
            frozen_items = [{"service_name": ci.service_name, "staff_id": ci.staff_id, "staff_name": ci.staff_name,
                             "total": ci.total, "commission_rate": ci.commission_rate, "commission_amount": ci.commission_amount}
                            for ci in ci_list]
            # Use first frozen rate if available
            for ci in ci_list:
                if ci.commission_rate is not None:
                    commission_rate = ci.commission_rate
                    break
            if ci_list:
                staff_name = ci_list[0].staff_name
        if not staff_name and inv.items:
            staff_name = inv.items[0].staff_name
            staff = db.query(Staff).filter(Staff.name == staff_name).first() if staff_name else None
            if staff:
                comm = db.query(StaffCommission).filter(StaffCommission.staff_id == staff.id).first()
                if comm:
                    commission_rate = comm.default_rate

        data["staff_commission_rate"] = commission_rate
        data["staff_name_primary"] = staff_name
        data["frozen_items"] = frozen_items  # Per-item frozen commissions
        result.append(data)

    if total_count is not None:
        return {"items": result, "total": total_count, "limit": limit, "offset": offset}
    return result


# ── DIAN POS (must be BEFORE /invoices/{invoice_id}) ──

@router.get("/invoices/pos-status")
def get_pos_status(db: Session = Depends(get_db), user=Depends(get_current_user)):
    tid = safe_tid(user, db)
    if not tid:
        raise HTTPException(400, "Tenant requerido")
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    if not tenant:
        raise HTTPException(404, "Tenant no encontrado")
    range_from = tenant.invoice_range_from or 1
    range_to = tenant.invoice_range_to or 4000
    total_range = range_to - range_from + 1
    used = db.query(func.count(Invoice.id)).filter(Invoice.tenant_id == tid, Invoice.is_pos == True, Invoice.dian_status != 'voided').scalar() or 0
    last_pos = db.query(func.max(Invoice.pos_number)).filter(Invoice.tenant_id == tid, Invoice.is_pos == True).scalar() or 0
    remaining = total_range - used
    prefix = tenant.invoice_prefix or 'POS'
    pending = db.query(func.count(Invoice.id)).filter(Invoice.tenant_id == tid, Invoice.is_pos == True, Invoice.dian_status == 'pending').scalar() or 0
    sent = db.query(func.count(Invoice.id)).filter(Invoice.tenant_id == tid, Invoice.is_pos == True, Invoice.dian_status.in_(['sent', 'accepted'])).scalar() or 0
    alerts = []
    if remaining <= 0:
        alerts.append("Se agoto el rango POS. Solicite nueva resolucion DIAN.")
    elif remaining <= 100:
        alerts.append(f"Quedan {remaining} facturas POS disponibles.")
    resolution_valid_to = tenant.resolution_valid_to
    if resolution_valid_to:
        days_left = (resolution_valid_to - date.today()).days
        if days_left <= 30 and days_left > 0:
            alerts.append(f"La resolucion vence en {days_left} dias.")
        elif days_left <= 0:
            alerts.append("La resolucion DIAN esta vencida.")
    fields = [tenant.nit, tenant.legal_name, tenant.dian_resolution_number, tenant.invoice_prefix, tenant.invoice_range_from, tenant.invoice_range_to]
    completeness = round(sum(1 for f in fields if f) / len(fields) * 100)
    return {"prefix": prefix, "range_from": range_from, "range_to": range_to, "total_range": total_range, "used": used, "remaining": remaining, "last_pos_number": last_pos, "pending": pending, "sent": sent, "completeness": completeness, "resolution_number": tenant.dian_resolution_number, "resolution_valid_to": resolution_valid_to.isoformat() if resolution_valid_to else None, "alerts": alerts, "nit": tenant.nit, "legal_name": tenant.legal_name}


@router.post("/invoices/assign-pos")
def assign_pos_numbers_v2(data: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    import logging; logger = logging.getLogger(__name__)
    tid = safe_tid(user, db)
    if not tid:
        raise HTTPException(400, "Tenant requerido")
    invoice_ids = data.get("invoice_ids", [])
    if not invoice_ids:
        raise HTTPException(400, "Seleccione al menos una factura")
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    if not tenant:
        raise HTTPException(404, "Tenant no encontrado")
    prefix = tenant.invoice_prefix or 'POS'
    range_from = tenant.invoice_range_from or 1
    range_to = tenant.invoice_range_to or 4000
    # Thread-safe: SELECT FOR UPDATE to prevent race condition on POS number assignment
    from sqlalchemy import text as sa_text
    result = db.execute(sa_text(
        "SELECT COALESCE(MAX(pos_number), :range_start) FROM invoice WHERE tenant_id = :tid AND is_pos = true FOR UPDATE"
    ), {"tid": tid, "range_start": range_from - 1})
    last_pos = result.scalar() or (range_from - 1)
    next_pos = last_pos + 1
    all_invs = db.query(Invoice).filter(Invoice.id.in_(invoice_ids), Invoice.tenant_id == tid).order_by(Invoice.issued_date, Invoice.id).all()
    invoices = [inv for inv in all_invs if not inv.is_pos]
    logger.info(f"[ASSIGN-POS] ids={invoice_ids}, found={len(all_invs)}, eligible={len(invoices)}")
    if not invoices:
        already = [f"{i.invoice_number}(is_pos={i.is_pos})" for i in all_invs]
        raise HTTPException(400, f"Ya tienen POS: {', '.join(already)}" if all_invs else f"No se encontraron facturas con IDs {invoice_ids}")
    if next_pos + len(invoices) - 1 > range_to:
        raise HTTPException(400, f"Rango insuficiente. Solo quedan {range_to - last_pos} consecutivos POS.")
    assigned = []
    for inv in invoices:
        inv.pos_number = next_pos
        inv.pos_prefix = prefix
        inv.pos_full_number = f"{prefix}-{str(next_pos).zfill(4)}"
        inv.is_pos = True
        inv.dian_status = 'pending'
        assigned.append({"id": inv.id, "invoice_number": inv.invoice_number, "pos_full_number": inv.pos_full_number, "pos_number": next_pos})
        next_pos += 1
    db.commit()
    return {"assigned": assigned, "count": len(assigned)}


@router.put("/invoices/{invoice_id}/void-pos")
def void_pos_invoice(invoice_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    tid = safe_tid(user, db)
    if not tid:
        raise HTTPException(400, "Tenant requerido")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.tenant_id == tid).first()
    if not inv:
        raise HTTPException(404, "Factura no encontrada")
    if not inv.is_pos:
        raise HTTPException(400, "Esta factura no tiene POS asignado")
    if inv.dian_status in ('sent', 'accepted'):
        raise HTTPException(400, "No se puede anular una factura ya enviada a la DIAN.")
    old_pos = inv.pos_full_number
    inv.pos_number = None
    inv.pos_prefix = None
    inv.pos_full_number = None
    inv.is_pos = False
    inv.dian_status = 'voided'
    db.commit()
    return {"message": f"Factura {old_pos} anulada", "invoice_id": invoice_id}


@router.post("/invoices/send-dian")
def send_invoices_to_dian(data: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Send pending POS invoices to DIAN via Alegra."""
    tid = safe_tid(user, db)
    if not tid:
        raise HTTPException(400, "Tenant requerido")
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    if not tenant:
        raise HTTPException(404, "Tenant no encontrado")

    # Check Alegra config
    alegra_email = getattr(tenant, 'fiscal_email', None) or getattr(tenant, 'owner_email', None)
    alegra_token = getattr(tenant, 'billing_provider_api_key', None)
    alegra_env = getattr(tenant, 'billing_environment', 'production')

    if not alegra_token:
        raise HTTPException(400, "Alegra no configurado. Ve a Configuracion > Facturacion y agrega tu API Token de Alegra.")

    from services.alegra import AlegraService
    alegra = AlegraService(email=alegra_email, token=alegra_token, environment=alegra_env)

    # Test connection
    ok, msg = alegra.test_connection()
    if not ok:
        raise HTTPException(400, f"Error conectando con Alegra: {msg}")

    invoice_ids = data.get("invoice_ids", [])
    if not invoice_ids:
        raise HTTPException(400, "Seleccione al menos una factura")

    invoices = db.query(Invoice).filter(
        Invoice.id.in_(invoice_ids),
        Invoice.tenant_id == tid,
        Invoice.is_pos == True,
        Invoice.dian_status == 'pending',
    ).all()

    if not invoices:
        raise HTTPException(400, "No hay facturas pendientes de envio a DIAN")

    results = []
    for inv in invoices:
        # Build invoice data for Alegra
        items_data = []
        for item in inv.items:
            items_data.append({
                "service_name": item.service_name,
                "quantity": item.quantity,
                "unit_price": item.unit_price,
                "total": item.total,
            })

        invoice_payload = {
            "client_name": inv.client_name,
            "client_document_type": inv.client_document_type or "CC",
            "client_document": inv.client_document,
            "client_email": inv.client_email,
            "client_phone": inv.client_phone,
            "client_address": getattr(inv, 'client_address', None),
            "items": items_data,
            "subtotal": inv.subtotal,
            "tax_rate": inv.tax_rate,
            "tax_amount": inv.tax_amount,
            "total": inv.total,
            "payment_method": inv.payment_method,
            "issued_date": inv.issued_date.isoformat() if inv.issued_date else date.today().isoformat(),
            "due_date": inv.due_date.isoformat() if inv.due_date else None,
            "payment_terms": inv.payment_terms,
            "notes": inv.notes,
            "discount_amount": inv.discount_amount,
        }

        result = alegra.create_and_stamp_invoice(invoice_payload)

        if result.get("error"):
            results.append({"id": inv.id, "invoice_number": inv.invoice_number, "success": False, "error": result.get("detail", "Error desconocido")})
            continue

        # Update invoice with Alegra response
        inv.alegra_id = result.get("alegra_id")
        inv.cufe = result.get("cufe")
        inv.dian_status = result.get("dian_status", "sent")
        inv.dian_sent_at = datetime.utcnow()
        inv.dian_response = result.get("raw_response", "")

        results.append({
            "id": inv.id,
            "invoice_number": inv.invoice_number,
            "pos_full_number": inv.pos_full_number,
            "success": True,
            "dian_status": inv.dian_status,
            "cufe": inv.cufe,
            "alegra_id": inv.alegra_id,
        })

    db.commit()
    sent_count = sum(1 for r in results if r.get("success"))
    failed_count = sum(1 for r in results if not r.get("success"))
    return {"results": results, "sent": sent_count, "failed": failed_count}


@router.get("/invoices/{invoice_id}")
def get_invoice(invoice_id: int, db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    """Single-invoice fetch — enriched with frozen commissions from CheckoutItem
    so the detail panel shows the exact rates that were locked at payment time
    (the same enrichment the list endpoint does)."""
    tid = safe_tid(current_user, db)
    q = db.query(Invoice).filter(Invoice.id == invoice_id)
    q = _tenant_filter(q, Invoice, tid)
    inv = q.first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")

    # Build the base response from the ORM model
    base = InvoiceResponse.model_validate(inv).model_dump()

    # ── Mirror the enrichment list_invoices applies ──
    from database.models import CheckoutItem, StaffServiceCommission
    checkout = db.query(Checkout).filter(Checkout.invoice_id == inv.id)
    if tid:
        checkout = checkout.filter(Checkout.tenant_id == tid)
    checkout = checkout.first()

    base["tip_amount"] = checkout.tip_amount if checkout and getattr(checkout, 'tip_amount', None) else 0
    base["receipt_url"] = checkout.receipt_url if checkout and getattr(checkout, 'receipt_url', None) else None

    commission_rate = 0.5
    frozen_items = []
    staff_name_primary = None

    if checkout:
        ci_list = db.query(CheckoutItem).filter(CheckoutItem.checkout_id == checkout.id).all()
        # Build product-commission map by parsing <!--PRODUCTS:...:PRODUCTS--> JSON
        # from related visit notes. The pos endpoint writes one VisitHistory per
        # staff member when a checkout is created, with the products array in
        # `notes`. Match by (client_id + staff_id) and same visit_date as the
        # checkout — payment_id is only set later during nómina, so filtering
        # on it would miss every recent sale.
        import json as _json_pc
        # Per-staff map so each staff's products are scoped to that staff
        # → key: (staff_id, name_lower) → comm-per-unit
        product_comm_map = {}
        try:
            from database.models import VisitHistory
            staff_ids_in_checkout = list({ci.staff_id for ci in ci_list if ci.staff_id})
            checkout_date = checkout.created_at.date() if checkout.created_at else None
            v_q = db.query(VisitHistory).filter(VisitHistory.notes.contains('<!--PRODUCTS:'))
            if checkout.client_id:
                v_q = v_q.filter(VisitHistory.client_id == checkout.client_id)
            if staff_ids_in_checkout:
                v_q = v_q.filter(VisitHistory.staff_id.in_(staff_ids_in_checkout))
            if checkout_date:
                v_q = v_q.filter(VisitHistory.visit_date == checkout_date)
            for v in v_q.all():
                if not v.notes or '<!--PRODUCTS:' not in v.notes:
                    continue
                try:
                    ps = v.notes.index('<!--PRODUCTS:') + len('<!--PRODUCTS:')
                    pe = v.notes.index(':PRODUCTS-->')
                    for p in _json_pc.loads(v.notes[ps:pe]):
                        name = (p.get('name') or '').strip().lower()
                        qty = max(int(p.get('qty') or 1), 1)
                        comm_total = int(p.get('comm') or 0)
                        if name and comm_total:
                            key = (v.staff_id, name)
                            product_comm_map[key] = max(product_comm_map.get(key, 0), comm_total // qty)
                except Exception:
                    pass
        except Exception:
            pass

        frozen_items = []
        for ci in ci_list:
            ci_amount = ci.commission_amount
            ci_rate = ci.commission_rate
            # Product line — fixed $ per unit, snapshot lives in visit notes
            if (ci.service_name or '').lower().startswith('[producto]') and ci_amount in (None, 0):
                clean_name = ci.service_name.replace('[Producto] ', '').replace('[producto] ', '').strip().lower()
                per_unit = product_comm_map.get((ci.staff_id, clean_name), 0)
                if per_unit == 0:
                    # Fallback: any staff with that product name
                    for (_sid, _name), v in product_comm_map.items():
                        if _name == clean_name and v > per_unit:
                            per_unit = v
                if per_unit > 0:
                    ci_amount = per_unit * (ci.quantity or 1)
            frozen_items.append({
                "service_id": getattr(ci, 'service_id', None),
                "service_name": ci.service_name,
                "staff_id": ci.staff_id,
                "staff_name": ci.staff_name,
                "total": ci.total,
                "commission_rate": ci_rate,
                "commission_amount": ci_amount,
            })
        for ci in ci_list:
            if ci.commission_rate is not None:
                commission_rate = ci.commission_rate
                break
        if ci_list:
            staff_name_primary = ci_list[0].staff_name

    if not staff_name_primary and inv.items:
        staff_name_primary = inv.items[0].staff_name
        st = db.query(Staff).filter(Staff.name == staff_name_primary).first() if staff_name_primary else None
        if st:
            comm = db.query(StaffCommission).filter(StaffCommission.staff_id == st.id).first()
            if comm:
                commission_rate = comm.default_rate

    base["staff_commission_rate"] = commission_rate
    base["staff_name_primary"] = staff_name_primary
    base["frozen_items"] = frozen_items

    return base


@router.post("/invoices/", response_model=InvoiceResponse)
def create_invoice(data: InvoiceCreate, db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    from sqlalchemy import text as sa_text
    tid = safe_tid(current_user, db)

    # Thread-safe invoice number: use SELECT FOR UPDATE to lock
    result = db.execute(sa_text(
        "SELECT invoice_number FROM invoice WHERE tenant_id = :tid "
        "ORDER BY id DESC LIMIT 1 FOR UPDATE"
    ), {"tid": tid or 0})
    row = result.fetchone()
    if row and row[0]:
        try:
            num = int(row[0].replace("FV-", ""))
            next_num = num + 1
        except (ValueError, AttributeError):
            next_num = 1
    else:
        next_num = 1
    invoice_number = f"FV-{next_num:04d}"

    # Calculate totals with discount
    subtotal = sum(item.unit_price * item.quantity for item in data.items)

    discount_amount = 0
    if data.discount_type == 'percent' and data.discount_value > 0:
        discount_amount = round(subtotal * data.discount_value / 100)
    elif data.discount_type == 'fixed' and data.discount_value > 0:
        discount_amount = min(data.discount_value, subtotal)

    # IVA INCLUIDO: price already has IVA, decompose for reporting
    taxable = subtotal - discount_amount
    if data.tax_rate > 0:
        base_amount = round(taxable / (1 + data.tax_rate))
        tax_amount = taxable - base_amount
        total = taxable  # Client pays same amount, IVA is internal
    else:
        base_amount = taxable
        tax_amount = 0
        total = taxable

    inv = Invoice(
        tenant_id=tid,
        invoice_number=invoice_number,
        client_id=data.client_id,
        client_name=data.client_name,
        client_phone=data.client_phone,
        client_document=data.client_document,
        client_document_type=data.client_document_type,
        client_email=data.client_email,
        client_address=data.client_address,
        subtotal=base_amount if data.tax_rate > 0 else subtotal,
        discount_type=data.discount_type,
        discount_value=data.discount_value,
        discount_amount=discount_amount,
        tax_rate=data.tax_rate,
        tax_amount=tax_amount,
        total=total,
        payment_method=data.payment_method,
        payment_terms=data.payment_terms or "contado",
        due_date=data.due_date,
        status="draft",
        issued_date=data.issued_date or datetime.utcnow().date(),
        notes=data.notes,
    )
    db.add(inv)
    db.flush()

    for item_data in data.items:
        item = InvoiceItem(
            tenant_id=tid,
            invoice_id=inv.id,
            service_id=item_data.service_id,
            service_name=item_data.service_name,
            quantity=item_data.quantity,
            unit_price=item_data.unit_price,
            total=item_data.unit_price * item_data.quantity,
            staff_id=item_data.staff_id,
            staff_name=item_data.staff_name,
            visit_id=item_data.visit_id,
        )
        db.add(item)

        # Mark visit as invoiced
        if item_data.visit_id:
            visit = db.query(VisitHistory).filter(VisitHistory.id == item_data.visit_id).first()
            if visit:
                visit.is_invoiced = True

    db.commit()
    db.refresh(inv)
    return inv


@router.put("/invoices/{invoice_id}", response_model=InvoiceResponse)
def update_invoice(invoice_id: int, data: InvoiceUpdate, db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    tid = safe_tid(current_user, db)
    q = db.query(Invoice).filter(Invoice.id == invoice_id)
    q = _tenant_filter(q, Invoice, tid)
    inv = q.first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")

    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(inv, k, v)

    # If marking as paid, set paid_at
    if data.status == "paid" and not inv.paid_at:
        inv.paid_at = datetime.utcnow()

    # If cancelling, cascade to checkout/visit/appointment/stock
    if data.status == "cancelled":
        _cancel_invoice_cascade(inv, db, tid)

    db.commit()
    db.refresh(inv)
    return inv


@router.delete("/invoices/{invoice_id}")
def cancel_invoice(invoice_id: int, db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    tid = safe_tid(current_user, db)
    q = db.query(Invoice).filter(Invoice.id == invoice_id)
    q = _tenant_filter(q, Invoice, tid)
    inv = q.first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    inv.status = "cancelled"
    inv.deleted_at = datetime.utcnow()
    _cancel_invoice_cascade(inv, db, tid)
    db.commit()
    return {"ok": True}


def _cancel_invoice_cascade(inv, db, tid):
    """Cancel invoice triggers full cascade:
    1. Invoice → status='cancelled', deleted_at set
    2. Checkout → status='voided'
    3. Checkout.invoice → status='cancelled' (redundant but safe)
    4. VisitHistory → status='cancelled' (single + multi-staff via appointment)
    5. Appointment → status reverts from 'paid' to 'confirmed'
    6. Product stock → restored via reversal InventoryMovement
    """
    from routes.pos_endpoints import _void_checkout_cascade
    # Find checkout linked to this invoice
    checkout_q = db.query(Checkout).filter(Checkout.invoice_id == inv.id)
    if tid:
        checkout_q = checkout_q.filter(Checkout.tenant_id == tid)
    checkout = checkout_q.first()
    if checkout and checkout.status != "voided":
        _void_checkout_cascade(checkout, db, tid)


# ============================================================================
# P&L & PAYMENT METHODS
# ============================================================================

@router.get("/finances/pnl", response_model=PnLResponse)
def get_pnl(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    start, end = _parse_period(period, date_from, date_to)

    # Revenue
    q = (
        db.query(VisitHistory)
        .filter(VisitHistory.status == "completed")
        .filter(VisitHistory.visit_date >= start, VisitHistory.visit_date <= end)
    )
    q = _tenant_filter(q, VisitHistory, tid)
    visits = q.all()
    total_revenue = sum(v.amount for v in visits)

    # Expenses
    eq = db.query(Expense).filter(Expense.date >= start, Expense.date <= end)
    eq = _tenant_filter(eq, Expense, tid)
    expenses = eq.all()
    total_expenses = sum(e.amount for e in expenses)

    # Commissions — per-service rates (same logic as Nómina)
    # Build service catalog
    svc_cat_pnl = {}
    svc_q_pnl = db.query(Service)
    if tid:
        svc_q_pnl = svc_q_pnl.filter(Service.tenant_id == tid)
    for svc in svc_q_pnl.all():
        svc_cat_pnl[svc.name.lower().strip()] = svc.id

    # Load all commission configs
    all_ssc = {}
    for r in db.query(StaffServiceCommission).all():
        all_ssc[(r.staff_id, r.service_id)] = r.commission_rate

    all_default_rates = {}
    for c in db.query(StaffCommission).all():
        all_default_rates[c.staff_id] = c.default_rate

    total_commissions = 0
    total_tips = 0
    for v in visits:
        default_rate = all_default_rates.get(v.staff_id, 0.40)
        names = v.service_name.split(',') if v.service_name else ['Otros']
        per_svc = (v.amount or 0) / max(len(names), 1)
        for name in names:
            name_clean = name.strip()
            if name_clean.startswith('[Producto]'):
                continue  # Products: commission from notes, not %
            svc_id = svc_cat_pnl.get(name_clean.lower().strip())
            rate = all_ssc.get((v.staff_id, svc_id), default_rate) if svc_id else default_rate
            total_commissions += round(per_svc * rate)
        # Product commissions from notes
        if v.notes and '<!--PRODUCTS:' in v.notes:
            import json as _jpnl
            try:
                ps = v.notes.index('<!--PRODUCTS:') + len('<!--PRODUCTS:')
                pe = v.notes.index(':PRODUCTS-->')
                prods = _jpnl.loads(v.notes[ps:pe])
                total_commissions += sum(p.get('comm', 0) or 0 for p in prods)
            except Exception:
                pass
        total_tips += getattr(v, 'tip', 0) or 0

    # Fines (income for business)
    fines_q = db.query(StaffFine).filter(
        StaffFine.fine_date >= start,
        StaffFine.fine_date <= end,
    )
    if tid:
        fines_q = fines_q.filter(StaffFine.tenant_id == tid)
    total_fines = sum(f.amount for f in fines_q.all())

    net_profit = total_revenue - total_expenses - total_commissions - total_tips + total_fines
    margin_pct = round((net_profit / total_revenue * 100), 1) if total_revenue > 0 else 0.0

    return PnLResponse(
        period=period,
        date_from=start.isoformat(),
        date_to=end.isoformat(),
        total_revenue=total_revenue,
        total_expenses=total_expenses,
        total_commissions=total_commissions,
        total_tips=total_tips,
        total_fines=total_fines,
        net_profit=net_profit,
        margin_pct=margin_pct,
    )


@router.get("/finances/payment-methods")
def get_payment_methods(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    start, end = _parse_period(period, date_from, date_to)
    q = (
        db.query(VisitHistory)
        .filter(VisitHistory.status == "completed")
        .filter(VisitHistory.visit_date >= start, VisitHistory.visit_date <= end)
    )
    q = _tenant_filter(q, VisitHistory, tid)
    visits = q.all()

    method_map = {}
    for v in visits:
        m = v.payment_method or "Sin registrar"
        if m not in method_map:
            method_map[m] = {"count": 0, "total": 0}
        method_map[m]["count"] += 1
        method_map[m]["total"] += v.amount

    grand_total = sum(info["total"] for info in method_map.values()) or 1
    items = sorted(
        [
            PaymentMethodItem(
                method=m,
                count=info["count"],
                total=info["total"],
                pct_of_total=round((info["total"] / grand_total * 100), 1),
            )
            for m, info in method_map.items()
        ],
        key=lambda x: x.total,
        reverse=True,
    )
    return {"items": items}


# ============================================================================
# EXPORT / IMPORT CLIENTS
# ============================================================================

@router.get("/clients/export")
def export_clients(db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    tid = safe_tid(current_user, db)
    q = db.query(Client)
    q = _tenant_filter(q, Client, tid)
    clients = q.order_by(Client.name).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "Codigo", "Nombre", "Telefono", "Email", "Cumpleanos",
        "Estado", "Visitas", "Total Gastado", "Ticket Promedio",
        "Ultima Visita", "Profesional Preferido", "Tags",
    ])

    for c in clients:
        completed = [v for v in c.visits if v.status == "completed"]
        total_spent = sum(v.amount for v in completed)
        visit_count = len(completed)
        avg_ticket = total_spent // visit_count if visit_count > 0 else 0
        last_visit = max((v.visit_date for v in completed), default=None)
        pref_barber = c.preferred_barber.name if c.preferred_barber else ""
        tags_str = ", ".join(c.tags) if c.tags else ""
        birthday = c.birthday.isoformat() if c.birthday else ""
        status = c.status_override or ("activo" if c.is_active else "inactivo")

        writer.writerow([
            c.id, c.client_id, c.name, c.phone, c.email or "",
            birthday, status, visit_count, total_spent, avg_ticket,
            last_visit.isoformat() if last_visit else "", pref_barber, tags_str,
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=clientes.csv"},
    )


@router.post("/clients/import", response_model=ImportResult)
async def import_clients(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)):
    tid = safe_tid(current_user, db)

    if not file.filename:
        raise HTTPException(status_code=400, detail="Archivo requerido")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    content = await file.read()

    rows = []
    if ext == "csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers_row = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2):
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers_row):
                        row_dict[headers_row[i]] = str(cell.value or "").strip()
                rows.append(row_dict)
        except ImportError:
            raise HTTPException(status_code=400, detail="Instala openpyxl para importar Excel")
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado. Usa CSV o XLSX")

    imported = 0
    skipped = 0
    errors = []
    row_results: list[ImportRowResult] = []
    new_client_objs: list = []  # collected to attach batch_id after batch is created

    # Normalize header keys (case-insensitive lookup)
    def get_val(row, *keys):
        for k in keys:
            for rk in row:
                if rk.lower().strip() == k.lower():
                    return row[rk].strip()
        return ""

    # Get max client_id number (scoped to tenant)
    client_q = db.query(Client.client_id)
    client_q = _tenant_filter(client_q, Client, tid)
    existing_ids = {c.client_id for c in client_q.all()}
    max_num = 0
    for cid in existing_ids:
        try:
            num = int("".join(filter(str.isdigit, cid)))
            max_num = max(max_num, num)
        except ValueError:
            pass

    # Build phone → (client_id, name) map so duplicates can reference the original
    phones_q = db.query(Client.phone, Client.client_id, Client.name)
    phones_q = _tenant_filter(phones_q, Client, tid)
    existing_phone_map: dict[str, tuple[str, str]] = {
        p: (cid, nm) for p, cid, nm in phones_q.all() if p
    }

    for i, row in enumerate(rows):
        line_no = i + 2  # +1 for header, +1 because spreadsheet rows are 1-indexed
        name = get_val(row, "nombre", "name")
        phone = get_val(row, "telefono", "phone", "tel")
        email_raw = get_val(row, "email", "correo")
        bday_str = get_val(row, "cumpleanos", "cumpleaños", "birthday")

        try:
            if not name or not phone:
                reason = "Falta nombre" if not name else "Falta teléfono"
                errors.append(f"Fila {line_no}: {reason.lower()}")
                row_results.append(ImportRowResult(
                    line=line_no, status="error", name=name or None, phone=phone or None,
                    email=email_raw or None, birthday=bday_str or None, reason=reason,
                ))
                continue

            if phone in existing_phone_map:
                existing_id, existing_name = existing_phone_map[phone]
                skipped += 1
                row_results.append(ImportRowResult(
                    line=line_no, status="duplicate", name=name, phone=phone,
                    email=email_raw or None, birthday=bday_str or None,
                    reason="Teléfono ya registrado",
                    existing_client_id=existing_id,
                    existing_client_name=existing_name,
                ))
                continue

            max_num += 1
            client_id = f"M{max_num:05d}"
            while client_id in existing_ids:
                max_num += 1
                client_id = f"M{max_num:05d}"

            email = email_raw or None
            birthday = None
            if bday_str:
                try:
                    birthday = date.fromisoformat(bday_str)
                except ValueError:
                    pass

            # Tags are assigned automatically by our system based on behaviour —
            # never imported from the client's spreadsheet.
            client = Client(
                tenant_id=tid,
                client_id=client_id,
                name=name,
                phone=phone,
                email=email,
                birthday=birthday,
                tags=[],
                is_active=True,
                accepts_whatsapp=True,
            )
            db.add(client)
            new_client_objs.append(client)
            existing_ids.add(client_id)
            existing_phone_map[phone] = (client_id, name)
            imported += 1
            row_results.append(ImportRowResult(
                line=line_no, status="imported", name=name, phone=phone,
                email=email, birthday=bday_str or None, client_id=client_id,
            ))
        except Exception as e:
            msg = str(e)[:80]
            errors.append(f"Fila {line_no}: {msg}")
            row_results.append(ImportRowResult(
                line=line_no, status="error", name=name or None, phone=phone or None,
                reason=msg,
            ))

    # Always create a batch record (even with 0 imported) so the audit trail
    # shows the user attempted an import — this matters for traceability.
    batch_id = None
    if len(rows) > 0:
        from database.models import ImportBatch
        # snapshot of admin name at import time
        admin_obj = db.query(Admin).filter(Admin.id == current_user.id).first()
        admin_name = admin_obj.name if admin_obj and getattr(admin_obj, 'name', None) else (
            getattr(current_user, 'username', None) or 'Sistema'
        )
        # Capture skip/error reasons for audit (first 50)
        audit_log = []
        for r in row_results:
            if r.status != 'imported':
                msg = f"Fila {r.line} [{r.status}]: {r.reason or ''}"
                if r.existing_client_id:
                    msg += f" → ya pertenece a {r.existing_client_id}"
                audit_log.append(msg)
                if len(audit_log) >= 50:
                    break

        batch = ImportBatch(
            tenant_id=tid,
            admin_id=current_user.id,
            admin_name=admin_name,
            filename=file.filename,
            file_size=len(content),
            total_rows=len(rows),
            imported_count=imported,
            skipped_count=skipped,
            error_count=sum(1 for r in row_results if r.status == 'error'),
            error_log=audit_log,
        )
        db.add(batch)
        db.flush()  # populate batch.id
        batch_id = batch.id

        # Tag every newly-imported client with the batch id
        for c in new_client_objs:
            c.import_batch_id = batch_id

        db.commit()

    return ImportResult(
        imported=imported,
        skipped=skipped,
        errors=errors[:20],
        total=len(rows),
        rows=row_results,
        batch_id=batch_id,
    )


# ============================================================================
# IMPORT HISTORY
# ============================================================================
@router.get("/clients/import-history")
async def list_import_batches(
    limit: int = 30,
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    """List bulk-import batches for this tenant, most recent first."""
    from database.models import ImportBatch
    tid = safe_tid(current_user, db)

    q = db.query(ImportBatch).filter(ImportBatch.tenant_id == tid)\
        .order_by(ImportBatch.created_at.desc()).limit(min(limit, 100))

    batches = q.all()
    return [
        {
            "id": b.id,
            "filename": b.filename,
            "file_size": b.file_size,
            "total_rows": b.total_rows,
            "imported_count": b.imported_count,
            "skipped_count": b.skipped_count,
            "error_count": b.error_count,
            "admin_name": b.admin_name,
            "created_at": b.created_at.isoformat() if b.created_at else None,
        }
        for b in batches
    ]


@router.get("/clients/import-history/{batch_id}")
async def get_import_batch(
    batch_id: int,
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    """Get one import batch with the clients that came from it."""
    from database.models import ImportBatch
    tid = safe_tid(current_user, db)

    batch = db.query(ImportBatch).filter(
        ImportBatch.id == batch_id,
        ImportBatch.tenant_id == tid,
    ).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Importación no encontrada")

    clients_q = db.query(Client).filter(
        Client.import_batch_id == batch_id,
        Client.tenant_id == tid,
    ).order_by(Client.id.asc()).all()

    return {
        "id": batch.id,
        "filename": batch.filename,
        "file_size": batch.file_size,
        "total_rows": batch.total_rows,
        "imported_count": batch.imported_count,
        "skipped_count": batch.skipped_count,
        "error_count": batch.error_count,
        "admin_name": batch.admin_name,
        "created_at": batch.created_at.isoformat() if batch.created_at else None,
        "error_log": batch.error_log or [],
        "clients": [
            {
                "id": c.id,
                "client_id": c.client_id,
                "name": c.name,
                "phone": c.phone,
                "email": c.email,
                "is_active": c.is_active,
            }
            for c in clients_q
        ],
    }


# ============================================================================
# ANALYTICS
# ============================================================================

WEEKDAY_NAMES = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]


def _calc_change_pct(current, previous):
    if previous > 0:
        return round(((current - previous) / previous * 100), 1)
    return 100.0 if current > 0 else 0


def _calc_commissions(visits, db):
    staff_rev = {}
    for v in visits:
        if v.staff_id not in staff_rev:
            staff_rev[v.staff_id] = 0
        staff_rev[v.staff_id] += v.amount

    total = 0
    for sid, rev in staff_rev.items():
        comm = db.query(StaffCommission).filter(StaffCommission.staff_id == sid).first()
        rate = comm.default_rate if comm else 0.40
        total += int(rev * rate)
    return total


@router.get("/finances/analytics")
def get_finance_analytics(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    start, end = _parse_period(period, date_from, date_to)

    # Previous period
    period_days = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=period_days - 1)

    # Current period visits & expenses
    vq = (
        db.query(VisitHistory)
        .filter(VisitHistory.status == "completed")
        .filter(VisitHistory.visit_date >= start, VisitHistory.visit_date <= end)
    )
    vq = _tenant_filter(vq, VisitHistory, tid)
    visits = vq.all()

    eq = db.query(Expense).filter(Expense.date >= start, Expense.date <= end)
    eq = _tenant_filter(eq, Expense, tid)
    expenses = eq.all()

    # Previous period visits & expenses
    pvq = (
        db.query(VisitHistory)
        .filter(VisitHistory.status == "completed")
        .filter(VisitHistory.visit_date >= prev_start, VisitHistory.visit_date <= prev_end)
    )
    pvq = _tenant_filter(pvq, VisitHistory, tid)
    prev_visits = pvq.all()

    peq = db.query(Expense).filter(Expense.date >= prev_start, Expense.date <= prev_end)
    peq = _tenant_filter(peq, Expense, tid)
    prev_expenses = peq.all()

    total_revenue = sum(v.amount for v in visits)
    prev_revenue = sum(v.amount for v in prev_visits)
    total_expenses = sum(e.amount for e in expenses)
    prev_expenses_total = sum(e.amount for e in prev_expenses)
    total_commissions = _calc_commissions(visits, db)
    prev_commissions = _calc_commissions(prev_visits, db)
    net_profit = total_revenue - total_expenses - total_commissions
    prev_profit = prev_revenue - prev_expenses_total - prev_commissions

    # Revenue by day
    day_map = {}
    for v in visits:
        d = v.visit_date.isoformat()
        if d not in day_map:
            day_map[d] = {"revenue": 0, "visits": 0}
        day_map[d]["revenue"] += v.amount
        day_map[d]["visits"] += 1
    revenue_by_day = sorted(
        [
            {
                "date": d,
                "revenue": info["revenue"],
                "visits": info["visits"],
                "avg_ticket": info["revenue"] // info["visits"] if info["visits"] else 0,
            }
            for d, info in day_map.items()
        ],
        key=lambda x: x["date"],
    )

    # Best day / busiest day
    best_day_date = None
    best_day_revenue = 0
    busiest_day_date = None
    busiest_day_visits = 0
    for entry in revenue_by_day:
        if entry["revenue"] > best_day_revenue:
            best_day_revenue = entry["revenue"]
            best_day_date = entry["date"]
        if entry["visits"] > busiest_day_visits:
            busiest_day_visits = entry["visits"]
            busiest_day_date = entry["date"]

    # Revenue by staff
    staff_map = {}
    for v in visits:
        if v.staff_id not in staff_map:
            staff_map[v.staff_id] = {"revenue": 0, "count": 0}
        staff_map[v.staff_id]["revenue"] += v.amount
        staff_map[v.staff_id]["count"] += 1
    revenue_by_staff = []
    for sid, info in staff_map.items():
        staff = db.query(Staff).filter(Staff.id == sid).first()
        revenue_by_staff.append({
            "staff_id": sid,
            "staff_name": staff.name if staff else "Desconocido",
            "revenue": info["revenue"],
            "count": info["count"],
            "avg_ticket": info["revenue"] // info["count"] if info["count"] else 0,
            "pct_of_total": round((info["revenue"] / total_revenue * 100), 1) if total_revenue > 0 else 0,
        })
    revenue_by_staff.sort(key=lambda x: x["revenue"], reverse=True)

    # Revenue by service (with category lookup)
    svc_map = {}
    for v in visits:
        sn = v.service_name or "Sin servicio"
        if sn not in svc_map:
            svc_map[sn] = {"revenue": 0, "count": 0}
        svc_map[sn]["revenue"] += v.amount
        svc_map[sn]["count"] += 1
    revenue_by_service = []
    for sn, info in svc_map.items():
        svc = db.query(Service).filter(Service.name == sn).first()
        category = svc.category if svc else "Otros"
        revenue_by_service.append({
            "service_name": sn,
            "category": category,
            "revenue": info["revenue"],
            "count": info["count"],
            "pct_of_total": round((info["revenue"] / total_revenue * 100), 1) if total_revenue > 0 else 0,
        })
    revenue_by_service.sort(key=lambda x: x["revenue"], reverse=True)

    # Revenue by category
    cat_map = {}
    for item in revenue_by_service:
        cat = item["category"]
        if cat not in cat_map:
            cat_map[cat] = {"revenue": 0, "count": 0}
        cat_map[cat]["revenue"] += item["revenue"]
        cat_map[cat]["count"] += item["count"]
    revenue_by_category = sorted(
        [
            {
                "category": cat,
                "revenue": info["revenue"],
                "count": info["count"],
                "pct_of_total": round((info["revenue"] / total_revenue * 100), 1) if total_revenue > 0 else 0,
            }
            for cat, info in cat_map.items()
        ],
        key=lambda x: x["revenue"],
        reverse=True,
    )

    # Revenue by weekday
    wd_map = {}
    for v in visits:
        wd = v.visit_date.weekday()
        if wd not in wd_map:
            wd_map[wd] = {"revenue": 0, "visits": 0}
        wd_map[wd]["revenue"] += v.amount
        wd_map[wd]["visits"] += 1
    revenue_by_weekday = sorted(
        [
            {
                "weekday": wd,
                "weekday_name": WEEKDAY_NAMES[wd],
                "revenue": info["revenue"],
                "visits": info["visits"],
            }
            for wd, info in wd_map.items()
        ],
        key=lambda x: x["weekday"],
    )

    return {
        # Comparison
        "current_revenue": total_revenue,
        "previous_revenue": prev_revenue,
        "revenue_change_pct": _calc_change_pct(total_revenue, prev_revenue),
        "current_expenses": total_expenses,
        "previous_expenses": prev_expenses_total,
        "expenses_change_pct": _calc_change_pct(total_expenses, prev_expenses_total),
        "current_profit": net_profit,
        "previous_profit": prev_profit,
        "profit_change_pct": _calc_change_pct(net_profit, prev_profit),
        "current_visits": len(visits),
        "previous_visits": len(prev_visits),
        "visits_change_pct": _calc_change_pct(len(visits), len(prev_visits)),
        "current_avg_ticket": total_revenue // len(visits) if visits else 0,
        "previous_avg_ticket": prev_revenue // len(prev_visits) if prev_visits else 0,
        # Breakdowns
        "revenue_by_day": revenue_by_day,
        "revenue_by_staff": revenue_by_staff,
        "revenue_by_service": revenue_by_service,
        "revenue_by_category": revenue_by_category,
        "revenue_by_weekday": revenue_by_weekday,
        # Insights
        "best_day_date": best_day_date,
        "best_day_revenue": best_day_revenue,
        "busiest_day_date": busiest_day_date,
        "busiest_day_visits": busiest_day_visits,
        "unique_clients": len(set(v.client_id for v in visits if v.client_id)),
        "avg_ticket": total_revenue // len(visits) if visits else 0,
        "total_revenue": total_revenue,
        "total_visits": len(visits),
        "total_expenses": total_expenses,
        "total_commissions": total_commissions,
        "net_profit": net_profit,
        "margin_pct": round((net_profit / total_revenue * 100), 1) if total_revenue > 0 else 0,
    }


# ============================================================================
# EXPORT TRANSACTIONS
# ============================================================================
# STAFF PERFORMANCE — Detailed metrics per staff
# ============================================================================

@router.get("/finances/staff-performance")
def get_staff_performance(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    today = date.today()

    if date_from and date_to:
        start = date.fromisoformat(date_from)
        end = date.fromisoformat(date_to)
    elif period == "today":
        start = end = today
    elif period == "week":
        start = today - timedelta(days=today.weekday())
        end = today
    elif period == "year":
        start = date(today.year, 1, 1)
        end = today
    else:
        start = date(today.year, today.month, 1)
        end = today

    # Previous period
    days = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=days - 1)

    # All staff
    staff_q = db.query(Staff).filter(Staff.is_active == True)
    if tid:
        staff_q = staff_q.filter(Staff.tenant_id == tid)
    all_staff = staff_q.all()

    # Service catalog for categories + full objects for assignment lookup
    svc_catalog = {}
    svc_objects = {}  # {service_id: Service object}
    svc_q = db.query(Service)
    if tid:
        svc_q = svc_q.filter(Service.tenant_id == tid)
    for s in svc_q.all():
        svc_catalog[s.name.lower().strip()] = {"category": s.category or "Otros", "id": s.id}
        svc_objects[s.id] = s

    result = []
    for s in all_staff:
        # Current period visits
        vq = db.query(VisitHistory).filter(
            VisitHistory.staff_id == s.id,
            VisitHistory.status == "completed",
            VisitHistory.visit_date >= start,
            VisitHistory.visit_date <= end,
        )
        if tid:
            vq = vq.filter(VisitHistory.tenant_id == tid)
        visits = vq.all()

        # Previous period visits
        pvq = db.query(VisitHistory).filter(
            VisitHistory.staff_id == s.id,
            VisitHistory.status == "completed",
            VisitHistory.visit_date >= prev_start,
            VisitHistory.visit_date <= prev_end,
        )
        if tid:
            pvq = pvq.filter(VisitHistory.tenant_id == tid)
        prev_visits = pvq.all()

        # Commission — per-service from StaffServiceCommission
        comm = db.query(StaffCommission).filter(StaffCommission.staff_id == s.id).first()
        default_rate = comm.default_rate if comm else 0.40

        # Load all per-service commission rates for this staff
        ssc_rows = db.query(StaffServiceCommission).filter(
            StaffServiceCommission.staff_id == s.id
        ).all()
        ssc_map = {r.service_id: r.commission_rate for r in ssc_rows}  # {service_id: rate}

        # Metrics
        revenue = sum(v.amount or 0 for v in visits)
        prev_revenue = sum(v.amount or 0 for v in prev_visits)
        services_count = len(visits)
        prev_services = len(prev_visits)
        unique_clients = len(set(v.client_id for v in visits if v.client_id))
        avg_ticket = round(revenue / services_count) if services_count > 0 else 0

        # Build full service assignment list for this staff
        # All services where this staff is in staff_ids OR has a StaffServiceCommission
        assigned_services = {}  # {service_id: {name, price, rate}}
        for svc_name_lower, svc_data in svc_catalog.items():
            sid = svc_data.get("id")
            if not sid:
                continue
            # Check if staff is assigned to this service
            svc_obj = svc_objects.get(sid)
            if not svc_obj:
                continue
            staff_ids_list = svc_obj.staff_ids or []
            if s.id in staff_ids_list or sid in ssc_map:
                rate = ssc_map.get(sid, default_rate)
                assigned_services[sid] = {
                    "name": svc_obj.name,
                    "price": svc_obj.price or 0,
                    "rate": rate,
                    "count": 0,
                    "revenue": 0,
                    "commission": 0,
                }

        # Per-service commission breakdown from actual visits
        commission_breakdown_extra = {}  # for services done but not in assigned list
        commission_amount = 0
        for v in visits:
            names = v.service_name.split(',') if v.service_name else ['Otros']
            per_svc_amount = (v.amount or 0) / max(len(names), 1)
            for name in names:
                name = name.strip()
                is_product = name.startswith('[Producto]')
                svc_info = svc_catalog.get(name.lower().strip(), {})
                svc_id = svc_info.get("id")
                if is_product:
                    rate = 0
                elif svc_id and svc_id in ssc_map:
                    rate = ssc_map[svc_id]
                else:
                    rate = default_rate
                svc_comm = round(per_svc_amount * rate)
                commission_amount += svc_comm
                # Add to assigned_services if exists, else to extras
                if svc_id and svc_id in assigned_services:
                    assigned_services[svc_id]["count"] += 1
                    assigned_services[svc_id]["revenue"] += int(per_svc_amount)
                    assigned_services[svc_id]["commission"] += svc_comm
                else:
                    key = name
                    if key not in commission_breakdown_extra:
                        commission_breakdown_extra[key] = {"name": name, "price": 0, "rate": rate, "count": 0, "revenue": 0, "commission": 0}
                    commission_breakdown_extra[key]["count"] += 1
                    commission_breakdown_extra[key]["revenue"] += int(per_svc_amount)
                    commission_breakdown_extra[key]["commission"] += svc_comm

        # Extract real product commissions from PRODUCTS JSON (comm field)
        product_commissions_total = 0
        for v in visits:
            if v.notes and '<!--PRODUCTS:' in v.notes:
                import json as _json_pc
                try:
                    ps_pc = v.notes.index('<!--PRODUCTS:') + len('<!--PRODUCTS:')
                    pe_pc = v.notes.index(':PRODUCTS-->')
                    prods_pc = _json_pc.loads(v.notes[ps_pc:pe_pc])
                    product_commissions_total += sum(p.get('comm', 0) or 0 for p in prods_pc)
                except Exception:
                    pass
        commission_amount += product_commissions_total

        # Merge: assigned first (sorted by commission desc), then extras
        commission_breakdown = sorted(assigned_services.values(), key=lambda x: -x["commission"])
        commission_breakdown += sorted(commission_breakdown_extra.values(), key=lambda x: -x["commission"])

        # Fines in this period
        fines_q = db.query(StaffFine).filter(
            StaffFine.staff_id == s.id,
            StaffFine.fine_date >= start,
            StaffFine.fine_date <= end,
            StaffFine.is_paid == False,
        )
        if tid:
            fines_q = fines_q.filter(StaffFine.tenant_id == tid)
        fines = fines_q.all()
        fines_total = sum(f.amount for f in fines)
        fines_count = len(fines)

        # Revenue by service category
        cat_breakdown = {}
        for v in visits:
            names = v.service_name.split(',') if v.service_name else ['Otros']
            per_service = (v.amount or 0) / max(len(names), 1)
            for name in names:
                name = name.strip()
                if name.startswith('[Producto]'):
                    cat = 'Productos'
                else:
                    cat = svc_catalog.get(name.lower().strip(), {}).get("category", "Otros")
                cat_breakdown[cat] = cat_breakdown.get(cat, 0) + int(per_service)

        # Revenue by day
        day_revenue = {}
        for v in visits:
            d = v.visit_date.isoformat()
            day_revenue[d] = day_revenue.get(d, 0) + (v.amount or 0)

        # Top services
        svc_counts = {}
        for v in visits:
            for name in (v.service_name or '').split(','):
                name = name.strip()
                if name and not name.startswith('[Producto]'):
                    svc_counts[name] = svc_counts.get(name, 0) + 1
        top_services = sorted(svc_counts.items(), key=lambda x: -x[1])[:5]

        # Growth
        rev_growth = round(((revenue - prev_revenue) / prev_revenue) * 100, 1) if prev_revenue > 0 else (100.0 if revenue > 0 else 0)
        svc_growth = round(((services_count - prev_services) / prev_services) * 100, 1) if prev_services > 0 else (100.0 if services_count > 0 else 0)

        # Best day
        best_day = max(day_revenue.items(), key=lambda x: x[1]) if day_revenue else (None, 0)

        # Last visit (most recent in period)
        last_visit_data = None
        if visits:
            last_v = max(visits, key=lambda v: (v.visit_date, v.id))
            last_client = db.query(Client).filter(Client.id == last_v.client_id).first() if last_v.client_id else None
            last_visit_data = {
                "client_name": last_client.name if last_client else "Sin cliente",
                "service_name": last_v.service_name or "Servicio",
                "amount": last_v.amount or 0,
                "date": last_v.visit_date.isoformat(),
            }

        result.append({
            "staff_id": s.id,
            "staff_name": s.name,
            "staff_role": s.role or "",
            "photo_url": getattr(s, 'photo_url', None),
            "commission_rate": default_rate,
            "revenue": revenue,
            "prev_revenue": prev_revenue,
            "revenue_growth": rev_growth,
            "services_count": services_count,
            "prev_services": prev_services,
            "services_growth": svc_growth,
            "unique_clients": unique_clients,
            "avg_ticket": avg_ticket,
            "commission_amount": commission_amount,
            "commission_breakdown": [
                {"service": cb["name"], "price": cb["price"], "revenue": cb["revenue"], "rate": cb["rate"], "commission": cb["commission"], "count": cb["count"]}
                for cb in commission_breakdown
            ],
            "fines_total": fines_total,
            "fines_count": fines_count,
            "fines": [
                {"id": f.id, "reason": f.reason, "amount": f.amount, "date": f.fine_date.isoformat(), "notes": f.notes or ""}
                for f in fines
            ],
            "category_breakdown": cat_breakdown,
            "top_services": [{"name": n, "count": c} for n, c in top_services],
            "daily_revenue": [{"date": d, "revenue": r} for d, r in sorted(day_revenue.items())],
            "best_day": {"date": best_day[0], "revenue": best_day[1]} if best_day[0] else None,
            "last_visit": last_visit_data,
        })

    return sorted(result, key=lambda x: -x["revenue"])


@router.get("/finances/export")
def export_transactions(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    start, end = _parse_period(period, date_from, date_to)
    q = (
        db.query(VisitHistory)
        .filter(VisitHistory.status == "completed")
        .filter(VisitHistory.visit_date >= start, VisitHistory.visit_date <= end)
    )
    q = _tenant_filter(q, VisitHistory, tid)
    visits = q.order_by(VisitHistory.visit_date.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Fecha", "Cliente", "Servicio", "Profesional", "Monto", "Metodo de Pago"])

    for v in visits:
        client = db.query(Client).filter(Client.id == v.client_id).first()
        staff = db.query(Staff).filter(Staff.id == v.staff_id).first()
        writer.writerow([
            v.visit_date.isoformat(),
            client.name if client else "Desconocido",
            v.service_name or "",
            staff.name if staff else "Desconocido",
            v.amount,
            v.payment_method or "Sin registrar",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=transacciones_{start}_{end}.csv"},
    )


@router.get("/finances/export-excel")
def export_finance_excel(
    period: str = Query("month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    """Export multi-sheet Excel financial report (Resumen, Ingresos, Gastos, Comisiones, Nomina)."""
    tid = safe_tid(current_user, db)
    start, end = _parse_period(period, date_from, date_to)

    from services.reports.excel_export import generate_finance_report
    buffer = generate_finance_report(db, tid, start, end)

    filename = f"Finanzas_{start}_{end}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/clients/export-excel")
def export_clients_excel(
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    """Export all clients as Excel with full data."""
    tid = safe_tid(current_user, db)

    from services.reports.excel_export import generate_clients_report
    buffer = generate_clients_report(db, tid)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Clientes.xlsx"},
    )


# ============================================================================
# REVENUE FORECAST — Projection based on confirmed appointments + history
# ============================================================================

@router.get("/finances/forecast")
def revenue_forecast(
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    """Revenue forecast: confirmed appointments + historical averages per day-of-week."""
    from database.models import Appointment
    from routes._helpers import now_colombia

    tid = safe_tid(current_user, db)
    today = now_colombia().date()
    current_weekday = today.weekday()  # 0=Monday

    # ── 1. This week: confirmed appointment revenue ──
    week_start = today - timedelta(days=current_weekday)
    week_end = week_start + timedelta(days=6)

    week_apts_q = db.query(Appointment).filter(
        Appointment.date >= week_start,
        Appointment.date <= week_end,
        Appointment.status.in_(["confirmed", "completed", "paid"]),
    )
    if tid:
        week_apts_q = week_apts_q.filter(Appointment.tenant_id == tid)
    week_apts = week_apts_q.all()

    # Group by date
    week_confirmed = {}
    for a in week_apts:
        d = a.date.isoformat() if hasattr(a.date, 'isoformat') else str(a.date)
        week_confirmed[d] = week_confirmed.get(d, 0) + (a.price or 0)

    # ── 2. Historical averages per day-of-week (last 12 weeks) ──
    twelve_weeks_ago = today - timedelta(weeks=12)
    hist_q = db.query(
        func.extract('dow', VisitHistory.visit_date).label('dow'),
        func.avg(VisitHistory.amount).label('avg_amount'),
        func.count(VisitHistory.id).label('visit_count'),
        func.sum(VisitHistory.amount).label('total_amount'),
    ).filter(
        VisitHistory.visit_date >= twelve_weeks_ago,
        VisitHistory.visit_date < today,
    )
    if tid:
        hist_q = hist_q.filter(VisitHistory.tenant_id == tid)
    hist_rows = hist_q.group_by('dow').all()

    # PostgreSQL dow: 0=Sunday, 1=Monday... Convert to Python weekday (0=Monday)
    hist_by_weekday = {}
    weeks_counted = {}
    for row in hist_rows:
        pg_dow = int(row.dow)  # 0=Sun, 1=Mon, ...
        py_weekday = (pg_dow - 1) % 7  # Convert to 0=Mon
        total = float(row.total_amount or 0)
        count = int(row.visit_count or 0)
        # Estimate number of weeks with data for this day
        hist_by_weekday[py_weekday] = total
        weeks_counted[py_weekday] = max(1, count // max(1, count // 12 if count > 12 else 1))

    # Calculate weekly average per day
    day_avg = {}
    for wd in range(7):
        total = hist_by_weekday.get(wd, 0)
        # Divide by ~12 weeks to get weekly average
        day_avg[wd] = round(total / 12, 0) if total > 0 else 0

    # ── 3. Build daily forecast for this week ──
    DIAS = ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom']
    daily_forecast = []
    total_confirmed = 0
    total_projected = 0

    for i in range(7):
        d = week_start + timedelta(days=i)
        d_iso = d.isoformat()
        weekday = d.weekday()
        confirmed = week_confirmed.get(d_iso, 0)
        historical_avg = day_avg.get(weekday, 0)
        is_past = d < today
        is_today_flag = d == today

        if is_past:
            # Past days: use actual revenue
            actual_q = db.query(func.coalesce(func.sum(VisitHistory.amount), 0)).filter(
                VisitHistory.visit_date == d,
            )
            if tid:
                actual_q = actual_q.filter(VisitHistory.tenant_id == tid)
            actual = float(actual_q.scalar() or 0)
            projected = actual
        else:
            # Future days: max of confirmed appointments or historical average
            actual = 0
            projected = max(confirmed, historical_avg)

        total_confirmed += confirmed
        total_projected += projected

        daily_forecast.append({
            "date": d_iso,
            "day_name": DIAS[weekday],
            "day_number": d.day,
            "is_past": is_past,
            "is_today": is_today_flag,
            "confirmed": round(confirmed, 0),
            "historical_avg": round(historical_avg, 0),
            "projected": round(projected, 0),
            "actual": round(actual, 0) if is_past else None,
        })

    # ── 4. Monthly projection ──
    month_start = date(today.year, today.month, 1)
    next_month = date(today.year + (1 if today.month == 12 else 0), (today.month % 12) + 1, 1)
    month_end = next_month - timedelta(days=1)
    days_in_month = (month_end - month_start).days + 1
    days_elapsed = (today - month_start).days + 1
    days_remaining = days_in_month - days_elapsed

    # Actual revenue this month so far
    month_actual_q = db.query(func.coalesce(func.sum(VisitHistory.amount), 0)).filter(
        VisitHistory.visit_date >= month_start,
        VisitHistory.visit_date <= today,
    )
    if tid:
        month_actual_q = month_actual_q.filter(VisitHistory.tenant_id == tid)
    month_actual = float(month_actual_q.scalar() or 0)

    # Last month total for comparison
    prev_month_start = date(today.year - (1 if today.month == 1 else 0), (today.month - 2) % 12 + 1, 1)
    prev_month_end = month_start - timedelta(days=1)
    prev_month_q = db.query(func.coalesce(func.sum(VisitHistory.amount), 0)).filter(
        VisitHistory.visit_date >= prev_month_start,
        VisitHistory.visit_date <= prev_month_end,
    )
    if tid:
        prev_month_q = prev_month_q.filter(VisitHistory.tenant_id == tid)
    prev_month_total = float(prev_month_q.scalar() or 0)

    # Daily run rate
    daily_run_rate = month_actual / max(1, days_elapsed)
    month_projected = month_actual + (daily_run_rate * days_remaining)

    # Trend vs last month
    if prev_month_total > 0:
        month_trend_pct = round(((month_projected - prev_month_total) / prev_month_total) * 100, 1)
    else:
        month_trend_pct = 0

    # ── 5. Confirmed appointments for rest of month ──
    future_confirmed_q = db.query(func.coalesce(func.sum(Appointment.price), 0)).filter(
        Appointment.date > today,
        Appointment.date <= month_end,
        Appointment.status == "confirmed",
    )
    if tid:
        future_confirmed_q = future_confirmed_q.filter(Appointment.tenant_id == tid)
    future_confirmed = float(future_confirmed_q.scalar() or 0)

    return {
        "week": {
            "start": week_start.isoformat(),
            "end": week_end.isoformat(),
            "daily": daily_forecast,
            "total_confirmed": round(total_confirmed, 0),
            "total_projected": round(total_projected, 0),
        },
        "month": {
            "name": today.strftime("%B %Y"),
            "actual_so_far": round(month_actual, 0),
            "projected_total": round(month_projected, 0),
            "daily_run_rate": round(daily_run_rate, 0),
            "days_elapsed": days_elapsed,
            "days_remaining": days_remaining,
            "prev_month_total": round(prev_month_total, 0),
            "trend_pct": month_trend_pct,
            "future_confirmed": round(future_confirmed, 0),
        },
    }


# ─── FINES / MULTAS ──────────────────────────────────────

@router.get("/finances/fines")
def list_fines(
    staff_id: Optional[int] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    q = db.query(StaffFine)
    if tid:
        q = q.filter(StaffFine.tenant_id == tid)
    if staff_id:
        q = q.filter(StaffFine.staff_id == staff_id)
    if date_from:
        q = q.filter(StaffFine.fine_date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(StaffFine.fine_date <= date.fromisoformat(date_to))

    fines = q.order_by(StaffFine.fine_date.desc()).all()
    # Enrich with staff name
    staff_names = {s.id: s.name for s in db.query(Staff).filter(Staff.tenant_id == tid).all()} if tid else {}
    return [
        {
            "id": f.id,
            "staff_id": f.staff_id,
            "staff_name": staff_names.get(f.staff_id, ""),
            "reason": f.reason,
            "amount": f.amount,
            "fine_date": f.fine_date.isoformat(),
            "notes": f.notes or "",
            "created_by": f.created_by or "",
            "created_at": f.created_at.isoformat() if f.created_at else None,
        }
        for f in fines
    ]


@router.post("/finances/fines")
def create_fine(
    data: dict,
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    if not tid:
        raise HTTPException(400, "Tenant requerido")

    staff_id = data.get("staff_id")
    reason = (data.get("reason") or "").strip()
    amount = data.get("amount", 0)
    fine_date_str = data.get("fine_date")

    if not staff_id or not reason or not amount:
        raise HTTPException(400, "staff_id, reason y amount son obligatorios")

    # Verify staff belongs to tenant
    staff = db.query(Staff).filter(Staff.id == staff_id, Staff.tenant_id == tid).first()
    if not staff:
        raise HTTPException(404, "Profesional no encontrado")

    fine = StaffFine(
        tenant_id=tid,
        staff_id=staff_id,
        reason=reason,
        amount=int(amount),
        fine_date=date.fromisoformat(fine_date_str) if fine_date_str else date.today(),
        notes=(data.get("notes") or "").strip(),
        created_by=current_user.username if hasattr(current_user, 'username') else str(current_user.id),
    )
    db.add(fine)
    db.commit()
    db.refresh(fine)
    return {
        "id": fine.id,
        "staff_id": fine.staff_id,
        "staff_name": staff.name,
        "reason": fine.reason,
        "amount": fine.amount,
        "fine_date": fine.fine_date.isoformat(),
        "notes": fine.notes or "",
        "created_by": fine.created_by or "",
        "created_at": fine.created_at.isoformat() if fine.created_at else None,
    }


@router.delete("/finances/fines/{fine_id}")
def delete_fine(
    fine_id: int,
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    tid = safe_tid(current_user, db)
    q = db.query(StaffFine).filter(StaffFine.id == fine_id)
    if tid:
        q = q.filter(StaffFine.tenant_id == tid)
    fine = q.first()
    if not fine:
        raise HTTPException(404, "Multa no encontrada")
    db.delete(fine)
    db.commit()
    return {"detail": "Multa eliminada"}


# ─── CASH REGISTER (Caja Registradora) ───────────────────

@router.get("/finances/cash-register")
def get_cash_register(
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    """Get current cash balance + recent movements."""
    tid = safe_tid(current_user, db)
    
    # Current balance = sum of all movements
    q = db.query(func.coalesce(func.sum(CashMovement.amount), 0)).filter()
    if tid:
        q = q.filter(CashMovement.tenant_id == tid)
    balance = int(q.scalar() or 0)
    
    # Today's movements
    from routes._helpers import now_colombia
    today = now_colombia().date()
    today_q = db.query(func.coalesce(func.sum(CashMovement.amount), 0)).filter(
        func.date(CashMovement.created_at) >= today,
    )
    if tid:
        today_q = today_q.filter(CashMovement.tenant_id == tid)
    today_total = int(today_q.scalar() or 0)
    
    # Recent movements (last 50)
    mvq = db.query(CashMovement)
    if tid:
        mvq = mvq.filter(CashMovement.tenant_id == tid)
    movements = mvq.order_by(CashMovement.created_at.desc()).limit(50).all()
    
    # Stats
    sales_today = 0
    deposits_today = 0
    withdrawals_today = 0
    for m in movements:
        if m.created_at and m.created_at.date() >= today:
            if m.movement_type == 'sale':
                sales_today += m.amount
            elif m.movement_type == 'deposit':
                deposits_today += m.amount
            elif m.movement_type == 'withdrawal':
                withdrawals_today += abs(m.amount)
    
    return {
        "balance": balance,
        "today_total": today_total,
        "sales_today": sales_today,
        "deposits_today": deposits_today,
        "withdrawals_today": withdrawals_today,
        "movements": [
            {
                "id": m.id,
                "type": m.movement_type,
                "amount": m.amount,
                "balance_after": m.balance_after,
                "description": m.description,
                "reference_type": m.reference_type,
                "reference_id": m.reference_id,
                "created_by": m.created_by,
                "created_at": m.created_at.isoformat() if m.created_at else None,
            }
            for m in movements
        ],
    }


@router.post("/finances/cash-register/movement")
def add_cash_movement(
    data: dict,
    db: Session = Depends(get_db),
    current_user: Admin = Depends(get_current_user),
):
    """Add money (deposit) or remove money (withdrawal) from cash register."""
    tid = safe_tid(current_user, db)
    if not tid:
        raise HTTPException(400, "Tenant requerido")
    
    movement_type = data.get("type", "deposit")  # deposit, withdrawal, adjustment
    amount = int(data.get("amount", 0))
    description = (data.get("description") or "").strip()
    
    if amount <= 0:
        raise HTTPException(400, "Monto debe ser mayor a 0")
    if not description:
        raise HTTPException(400, "Descripcion requerida")
    
    # Get current balance
    balance = int(db.query(func.coalesce(func.sum(CashMovement.amount), 0)).filter(
        CashMovement.tenant_id == tid
    ).scalar() or 0)
    
    # For withdrawals, amount is negative
    if movement_type == "withdrawal":
        actual_amount = -amount
    else:
        actual_amount = amount
    
    new_balance = balance + actual_amount
    
    mv = CashMovement(
        tenant_id=tid,
        movement_type=movement_type,
        amount=actual_amount,
        balance_after=new_balance,
        description=description,
        reference_type="manual",
        created_by=current_user.username if hasattr(current_user, 'username') else str(current_user.id),
    )
    db.add(mv)
    db.commit()
    db.refresh(mv)
    
    return {
        "id": mv.id,
        "type": mv.movement_type,
        "amount": mv.amount,
        "balance_after": mv.balance_after,
        "description": mv.description,
        "created_at": mv.created_at.isoformat() if mv.created_at else None,
    }


# (DIAN POS endpoints moved before /invoices/{invoice_id} to avoid route conflict)
